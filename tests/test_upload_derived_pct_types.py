"""T0 (1.6.0): the four pipeline-derived CDE percentages must be PLAIN FLOATS.

WHAT THIS CATCHES, AND IT WAS LIVE ON THE PUBLIC APP AT v1.5.7.

``upload_handler.load_uploaded_pipeline`` computes four CDE-level shares from
the per-project Y/N flags when the CDE leaves the corresponding CDE Profile
cell blank -- which is what the SHIPPED BLANK TEMPLATE'S OWN ROW 5 INSTRUCTS::

    * Fields marked with * can be left blank -- they will be computed
      automatically from the per-project flags in the Pipeline sheet
      (below_market_rate, unrelated_entity, us_territory, persistent_poverty).

The computation runs through pandas, so each share arrives as ``np.float64``,
not ``float``. ``streamlit_app.utils._scoring_attrs_only`` then filters blanks
with ``v not in ("", [], {}, None)``. ``in`` compares by equality, and
``np.float64(0.79) == []`` is not ``False`` -- it is an EMPTY ARRAY, whose
truth value numpy refuses to decide::

    ValueError: The truth value of an empty array is ambiguous.

Page 1 catches that in its ``except Exception`` and calls ``st.stop()``, so the
CDE sees "Failed to read file: The truth value of an empty array is ambiguous"
and the tool does nothing else. Following the template's own instruction made
the recommended path unusable.

WHY NO EXISTING GATE SAW IT. Every fixture that reaches ``_scoring_attrs_only``
either hand-writes a dict of Python literals or uploads
``pipeline_sample.xlsx``, whose starred cells are all FILLED -- so the derived
branch never runs and no numpy scalar is ever produced.
``tests/test_template_roundtrip.py`` does drive the derived branch, but reads
``cde_extra`` directly and never strips it.

THE FIX IS AT THE SOURCE, NOT AT THE FILTER. ``load_uploaded_pipeline`` is what
emits these values into a dict every other layer treats as plain Python, so it
is what coerces them. The filter is hardened too (see
``test_the_blank_filter_cannot_raise_on_a_numpy_scalar``), because a filter that
can be crashed by a value is a filter that will be crashed by the next one.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest

_REPO_ROOT = Path(__file__).resolve().parent.parent
_STREAMLIT_APP = _REPO_ROOT / "streamlit_app"
for _p in (str(_REPO_ROOT), str(_STREAMLIT_APP)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from nmtcapp.core.upload_handler import load_uploaded_pipeline  # noqa: E402

from tests.conftest import templates_dir  # noqa: E402

#: The four attrs upload_handler derives from per-project flags, and the
#: starred CDE Profile columns whose blankness triggers each derivation.
_DERIVED = {
    "products_below_market_pct": "Below-Market Rate Pct (0–1)*",
    "unrelated_entities_pct": "Unrelated Entities Pct (0–1)*",
    "pct_us_territories": "US Territories Pct (0–1)*",
    "pct_persistent_poverty": "Persistent Poverty Pct (0–1)*",
}


def _workbook_with_starred_cells_left_blank() -> bytes:
    """The shipped sample workbook, with the starred cells cleared.

    That is not a synthetic mutation: it is the state the blank template ships
    in and the state row 5 tells a CDE is acceptable. The three identity cells
    are overwritten because ``assert_not_sample_identity`` refuses the shipped
    fictional CDE.
    """
    wb = openpyxl.load_workbook(Path(templates_dir()) / "pipeline_sample.xlsx")
    ws = wb["CDE Profile"]
    ws.cell(row=4, column=1).value = "Cardinal Ridge Community Capital, LLC"
    ws.cell(row=4, column=2).value = "CDE-2020-0431"
    ws.cell(row=4, column=3).value = "84-3319027"
    headers = {str(ws.cell(3, c).value).strip(): c
               for c in range(1, ws.max_column + 1) if ws.cell(3, c).value}
    for label in _DERIVED.values():
        assert label in headers, (
            f"the shipped template no longer has a {label!r} column; this "
            "gate has lost its subject"
        )
        ws.cell(row=4, column=headers[label]).value = None
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture(scope="module")
def derived_extra() -> dict:
    _pipeline, cde_extra = load_uploaded_pipeline(
        _workbook_with_starred_cells_left_blank(), "cardinal_ridge.xlsx"
    )
    assert cde_extra is not None
    return cde_extra


@pytest.mark.parametrize("attr", sorted(_DERIVED))
def test_each_derived_share_is_a_plain_python_float(derived_extra, attr):
    value = derived_extra[attr]
    assert type(value) is float, (
        f"{attr} arrived as {type(value).__name__}, not float. Every layer "
        "downstream treats this dict as plain Python; a numpy scalar makes "
        "`v not in ('', [], {}, None)` raise instead of answering."
    )


def test_the_blank_filter_cannot_raise_on_a_numpy_scalar():
    """The filter is hardened too, independently of the source fix.

    Driven with an actual numpy scalar rather than the coerced value, so this
    keeps asking even if the source fix is later reverted or bypassed.
    """
    np = pytest.importorskip("numpy")
    from utils import _scoring_attrs_only

    out = _scoring_attrs_only(
        {"dbc_focus_years": 4, "products_below_market_pct": np.float64(0.79)},
        is_demo=False,
    )
    assert out["products_below_market_pct"] == pytest.approx(0.79)
    assert out["dbc_focus_years"] == 4


def test_a_template_instructed_blank_upload_reaches_a_scored_application():
    """End to end: blank starred cells must not stop the recommended path."""
    from utils import read_uploaded_cde_profile

    _pipeline, cde_extra = load_uploaded_pipeline(
        _workbook_with_starred_cells_left_blank(), "cardinal_ridge.xlsx"
    )
    parsed = read_uploaded_cde_profile(cde_extra, is_demo=False)
    for attr in _DERIVED:
        assert attr in parsed.scoring_attrs, (
            f"{attr} was derived from the pipeline flags and then dropped by "
            "the blank filter"
        )
