"""T3 (1.6.0): the two input paths must agree about what a CDE profile is.

``core.cde.REQUIRED_CDE_FIELDS`` is THE ONE LIST -- eight fields, derived from
``_FIELD_GUIDANCE``, and its own comment records what the third hand-maintained
copy cost: deleting ``governance`` from it passed the entire suite, so a
required field silently stopped being validated.

THE TWO PATHS DID NOT AGREE WITH IT, AND THE RECOMMENDED ONE WAS THE SHORT ONE:

    YAML  (``nmtcapp init`` -> cde_profile.yaml)   all eight
    xlsx  (CDE Profile sheet, ``_CDE_FIELD_MAP``)  six -- ``contact`` and
                                                   ``governance`` had NO
                                                   COLUMNS AT ALL

So the recommended path GUARANTEED ITS OWN INCOMPLETENESS. After 1.6.0 T1
restores the four the identity strip was throwing away, these two would have
been the permanent residue: every xlsx user, for ever, told its profile was
missing two fields the sheet never offered to collect.

THE RULING: ADD THE COLUMNS. A TEMPLATE-VERSION CHANGE, SAID OUT LOUD -- the
shipped workbooks' A1 title now reads v1.2, and this file gates that.

WHY ADD RATHER THAN STOP REQUIRING. Dropping the two on the xlsx path would
make the paths agree by WEAKENING THE DEFINITION, and it would make
``REQUIRED_CDE_FIELDS`` "the one list, except on the recommended path" -- the
exact drift shape the constant's own comment exists to prevent. Both fields
are real CDFI Fund application content: the Allocation Application asks for
governance composition and for a contact, and ``section_c_management`` already
has a governance table with rows for ``board_members`` and
``community_representatives`` that no xlsx upload could ever fill.

BACKWARD COMPATIBILITY IS NOT AT RISK AND IS ASSERTED BELOW.
``_parse_cde_profile_from_wb`` maps by HEADER TEXT and skips headers it does
not recognise, so a v1.1 workbook keeps loading and simply supplies neither
field -- which is exactly what it did before.

WHAT IS RECORDED RATHER THAN FIXED. ``ein``, ``headquarters_state`` and
``organization_type`` are collected by the sheet and have no ``CDEProfile``
attribute to land on. They are not in ``REQUIRED_CDE_FIELDS``, nothing reads
them, and inventing three attributes to hold values nothing renders is a
larger change than this release is scoped for. ``test_the_gap_between_the_two
_paths_is_recorded_not_hidden`` pins the list so it cannot grow unnoticed.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

_REPO_ROOT = Path(__file__).resolve().parent.parent
_STREAMLIT_APP = _REPO_ROOT / "streamlit_app"
for _p in (str(_REPO_ROOT), str(_STREAMLIT_APP)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from nmtcapp.core.cde import (  # noqa: E402
    CDE_FIELDS_WHERE_EMPTY_IS_AN_ANSWER,
    REQUIRED_CDE_FIELDS,
)
from nmtcapp.core.upload_handler import (  # noqa: E402
    _CDE_FIELD_MAP,
    CDE_PROFILE_COLUMNS_FOR_REQUIRED_FIELD,
)

from tests.conftest import templates_dir  # noqa: E402

_TEMPLATES = ("pipeline_template.xlsx", "pipeline_sample.xlsx")


def _cde_headers(name: str) -> set:
    wb = openpyxl.load_workbook(Path(templates_dir()) / name)
    ws = wb["CDE Profile"]
    return {str(ws.cell(3, c).value).strip()
            for c in range(1, ws.max_column + 1) if ws.cell(3, c).value is not None}


class TestTheOneListIsStillTheOneList:
    def test_the_map_is_keyed_by_the_one_list(self):
        """No fourth hand-maintained copy: this map's keys ARE the list."""
        covered = set(CDE_PROFILE_COLUMNS_FOR_REQUIRED_FIELD)
        expected = set(REQUIRED_CDE_FIELDS) - set(CDE_FIELDS_WHERE_EMPTY_IS_AN_ANSWER)
        assert covered == expected, (
            "CDE_PROFILE_COLUMNS_FOR_REQUIRED_FIELD has drifted from "
            "REQUIRED_CDE_FIELDS.\n"
            f"  only in the map:  {sorted(covered - expected)}\n"
            f"  only in the list: {sorted(expected - covered)}"
        )

    def test_prior_awards_is_the_documented_exception(self):
        """``[]`` is a complete answer, so no column is required for it."""
        assert "prior_awards" in CDE_FIELDS_WHERE_EMPTY_IS_AN_ANSWER
        assert "prior_awards" not in CDE_PROFILE_COLUMNS_FOR_REQUIRED_FIELD


class TestEveryRequiredFieldHasColumnsInTheShippedTemplates:
    @pytest.mark.parametrize("template", _TEMPLATES)
    @pytest.mark.parametrize("field", sorted(CDE_PROFILE_COLUMNS_FOR_REQUIRED_FIELD))
    def test_the_columns_exist(self, template, field):
        headers = _cde_headers(template)
        missing = [label for label in CDE_PROFILE_COLUMNS_FOR_REQUIRED_FIELD[field]
                   if label not in headers]
        assert not missing, (
            f"{template}'s CDE Profile sheet has no column for the required "
            f"field {field!r}: {missing}"
        )

    @pytest.mark.parametrize("field", sorted(CDE_PROFILE_COLUMNS_FOR_REQUIRED_FIELD))
    def test_every_named_column_is_actually_parsed(self, field):
        unmapped = [label for label in CDE_PROFILE_COLUMNS_FOR_REQUIRED_FIELD[field]
                    if label not in _CDE_FIELD_MAP]
        assert not unmapped, (
            f"these columns are named for {field!r} but _CDE_FIELD_MAP does "
            f"not parse them, so the sheet collects a value nothing reads: "
            f"{unmapped}"
        )


class TestTheShippedTemplatesDeclareTheNewVersion:
    @pytest.mark.parametrize("template", _TEMPLATES)
    def test_the_cde_profile_sheet_says_v1_2(self, template):
        wb = openpyxl.load_workbook(Path(templates_dir()) / template)
        title = str(wb["CDE Profile"].cell(1, 1).value or "")
        assert "v1.2" in title, (
            "the CDE Profile sheet gained columns and still calls itself "
            f"v1.1. Adding fields without a version is how a template "
            f"silently redefines itself. A1 reads: {title!r}"
        )


class TestAV11WorkbookStillLoads:
    def test_a_sheet_without_the_new_columns_still_parses(self):
        """Backward compatibility: v1.1 files keep working, supplying neither."""
        from nmtcapp.core.upload_handler import _parse_cde_profile_from_wb

        wb = openpyxl.load_workbook(Path(templates_dir()) / "pipeline_sample.xlsx")
        ws = wb["CDE Profile"]
        for c in range(1, ws.max_column + 1):
            header = str(ws.cell(3, c).value or "").strip()
            for labels in CDE_PROFILE_COLUMNS_FOR_REQUIRED_FIELD.values():
                if header in labels and header not in (
                    "CDE Name", "CDE ID", "Certification Date",
                    "Mission Statement", "Target Markets (states, comma-sep)",
                ):
                    ws.cell(3, c).value = None
                    ws.cell(4, c).value = None
        parsed = _parse_cde_profile_from_wb(wb)
        assert parsed, "a v1.1-shaped sheet stopped parsing entirely"
        assert "contact_name" not in parsed
        assert "governance_board_members" not in parsed


class TestTheRecommendedPathCanNowScoreCompleteness:
    """The end the whole ruling is for: an xlsx upload can be COMPLETE."""

    def test_a_fully_filled_sheet_passes_check_completeness(self):
        import io

        import streamlit as st

        from nmtcapp.core.application import Application
        from nmtcapp.core.upload_handler import load_uploaded_pipeline
        from nmtcapp.validation.completeness_check import check_completeness
        from utils import get_or_create_app, read_uploaded_cde_profile

        st.session_state.clear()
        wb = openpyxl.load_workbook(Path(templates_dir()) / "pipeline_sample.xlsx")
        ws = wb["CDE Profile"]
        headers = {str(ws.cell(3, c).value).strip(): c
                   for c in range(1, ws.max_column + 1) if ws.cell(3, c).value}
        for label, value in {
            "CDE Name": "Cardinal Ridge Community Capital, LLC",
            "CDE ID": "CDE-2020-0431",
            "EIN": "84-3319027",
            "Certification Date": "2020-09-14",
            "Mission Statement": "Deploy NMTC capital into distressed rural communities.",
            "Target Markets (states, comma-sep)": "North Carolina, Tennessee",
            "Contact Name": "A. Reyes",
            "Contact Email": "areyes@cardinalridgecapital.org",
            "Board Members": 7,
            "Community Representatives": 3,
        }.items():
            assert label in headers, f"no {label!r} column in the shipped template"
            ws.cell(4, headers[label]).value = value
        buf = io.BytesIO()
        wb.save(buf)

        pipeline, cde_extra = load_uploaded_pipeline(buf.getvalue(), "cr.xlsx")
        parsed = read_uploaded_cde_profile(cde_extra, is_demo=False)
        app = get_or_create_app(pipeline=pipeline, is_demo=False, cde_extra=parsed)
        result = check_completeness(app)
        cde_issues = [i for i in result.issues if i.startswith("CDE profile missing")]
        assert not cde_issues, (
            "a CDE Profile sheet filled in completely still reports missing "
            f"required fields: {cde_issues}"
        )
        st.session_state.clear()


def test_the_gap_between_the_two_paths_is_recorded_not_hidden():
    """Sheet columns with no ``CDEProfile`` attribute, pinned so they cannot grow."""
    from utils import _IDENTITY_TO_PROFILE_ATTR

    identity_keys = {
        "cde_name", "cde_id", "ein", "headquarters_state", "certification_date",
        "mission", "website", "organization_type", "target_markets",
        "contact_name", "contact_email",
        "governance_board_members", "governance_community_representatives",
    }
    lands_somewhere = (
        set(_IDENTITY_TO_PROFILE_ATTR)
        | {"target_markets", "contact_name", "contact_email",
           "governance_board_members", "governance_community_representatives"}
    )
    assert identity_keys - lands_somewhere == {
        "ein", "headquarters_state", "organization_type",
    }, (
        "the set of CDE Profile identity columns that reach nothing has "
        "changed. Three is the recorded state (see this module's docstring); "
        "a fourth is a new silent loss, and one fewer is a fix that should "
        "update this pin."
    )
