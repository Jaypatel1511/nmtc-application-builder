"""Every workbook label must fit the row it is drawn in (1.3.0 B1).

THE DEFECT THIS GATE EXISTS FOR
===============================

1.3.0 S4 added a denominator disclosure to the Summary Dashboard's distress
label and a basis note to carry the rest of the argument. The metrics loop set
``row_dimensions[row].height = 18`` for every row, uniformly, so the two-line
label kept a one-line height. Measured in Excel 16.112 at the shipped geometry:
A12 needed 30.0 pt and shipped 18.0. The visible text ended mid-negation —

    Deep/Severe Distress Concentration (a share of QEI, not

— and the half that never displayed was the pointer to the note. The round's
own remedy was invisible on the surface it was written for.

WHY THE RENDERED BASELINE COULD NOT CATCH IT, AND CANNOT PROVE THE FIX
======================================================================

tests/rendered_baseline/excel.txt records::

    Summary Dashboard!A12|str|fmt=General|<text>

Cell coordinates, Python types, number formats, values. Row heights, column
widths and merge ranges appear nowhere in it, so the entire class is invisible
to that gate — not under-tested, structurally unreachable. A correct fix to B1
leaves excel.txt byte-unchanged, which is why a byte-unchanged baseline is not
evidence here.

This is the same shape as the interpolation mask that hid every printed
constant until 1.2.1 built tests/pinned_constants.txt as a separate gate, and
the same shape as the value-only projection that hid B-3's number formats until
the excel_cell_formats surface was added. Third instance. The remedy is the
same one: a gate that reads the dimension the other gate cannot see.

WHAT THIS GATE ASSERTS

It loads the SHIPPED workbook and, for every text cell that carries a label,
re-derives the height that text needs from the file's own geometry — the actual
string, the actual merged span, the actual font size — and fails if the row is
shorter. It reads nothing from the builder, so a height hardcoded anywhere, a
column narrowed, a font enlarged, or a label lengthened all fail it.

It also asserts no label needs MORE than Excel's 409-pt row ceiling, because a
row cannot grow past that: text needing more is text that will clip however the
height is set, and has to be split across rows by a person.

THE GATE'S OWN SENSITIVITY IS TESTED. ``test_checker_catches_the_1_3_0_defect``
rebuilds B1 exactly — a two-line label at ``height = 18`` — and asserts the
checker flags it. A gate that cannot fail is the recurring defect of this
package; this one is made to fail on demand, in CI, next to the gate itself.
"""
from __future__ import annotations

import pytest

pytest.importorskip("openpyxl", reason="openpyxl not installed")

import openpyxl
from openpyxl.utils import get_column_letter

from nmtcapp.renderers._sheet_geometry import (
    DEFAULT_FONT_SIZE, DEFAULT_ROW_HEIGHT, MAX_ROW_HEIGHT, MIN_ROW_HEIGHT,
    chars_per_line, required_row_height, span_points, wrapped_line_count,
)
from nmtcapp.renderers.excel_builder import ExcelApplicationBuilder

#: EVERY sheet. This started as just the two hand-laid-out sheets, excluded
#: the six DataFrame ones with the reasoning "_write_df_to_sheet does not set
#: row heights at all, so those rows keep Excel's default and autofit applies
#: normally" — and that reasoning was FALSE. _write_df_to_sheet set
#: ``row_dimensions[row_num].height = 16`` for every data row, uniformly, the
#: same hardcoded height in the same shape as the metrics loop's 18, forty
#: lines above a column auto-sizer that had not run yet.
#:
#: Measured on the shipped sample before the fix: Investor Commitments!G6 held
#: the 351-character "[CDE TO COMPLETE: This table is a blank form ...]"
#: instruction in a 40-wide column at height 16 — one line of nine — and
#: Distress Documentation!N8 clipped the ACS vintage off "CDFI Fund NMTC
#: Eligibility Table (2016-2020 ACS)". The audit found B1 on one sheet; the
#: same defect was live on six more, and the first draft of this gate was
#: written to look away from them.
#:
#: So: no exclusions. A sheet this gate does not read is a sheet where the
#: next uniform height goes unnoticed.
LAID_OUT_SHEETS = None  # every sheet in the workbook


def _merged_span_points(ws, cell) -> float:
    """Width in points of the merged range anchored at ``cell``, or of its column."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row == cell.row and rng.min_col == cell.column:
            widths = []
            for col in range(rng.min_col, rng.max_col + 1):
                dim = ws.column_dimensions[get_column_letter(col)]
                # openpyxl reports None for a column never explicitly sized;
                # Excel's default is 8.43 characters.
                widths.append(dim.width if dim.width else 8.43)
            return span_points(*widths)
    dim = ws.column_dimensions[get_column_letter(cell.column)]
    return span_points(dim.width if dim.width else 8.43)


def _is_merged(ws, cell) -> bool:
    """Whether ``cell`` anchors a merged range on this sheet."""
    return any(rng.min_row == cell.row and rng.min_col == cell.column
               for rng in ws.merged_cells.ranges)


def check_sheet_geometry(ws) -> tuple:
    """Return ``(findings, checked, skipped)`` for one worksheet.

    ``checked`` and ``skipped`` are returned, not discarded, because
    ``assert not findings`` is true of a sheet this function looked at and
    found nothing wrong AND of a sheet it never looked at. Those are different
    facts and the gate has to be able to tell them apart — see
    ``test_the_gate_measures_every_text_cell_in_the_workbook``.

    Example::

        findings, checked, skipped = check_sheet_geometry(wb["Summary Dashboard"])
        assert not findings and checked and not skipped
    """
    findings = []
    checked = 0
    skipped = []
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or not cell.value.strip():
                continue
            font_size = float(cell.font.size or 11)
            span = _merged_span_points(ws, cell)
            needed = required_row_height(cell.value, span, font_size)
            shipped = ws.row_dimensions[cell.row].height
            merged = _is_merged(ws, cell)
            if not shipped:
                # An unset height on an UNMERGED row is Excel's instruction to
                # autofit on render; such a row cannot clip and is not this
                # gate's business.
                #
                # ON A MERGED RANGE IT IS NOT. This module's own docstring says
                # so in capitals — "MERGED CELLS DO NOT AUTOFIT ... Excel's
                # AutoFit is a no-op on a merged range" — and the skip was
                # written on the opposite premise, so every merged label with
                # no explicit height was waved through by a rule the module
                # contradicts. Such a row is drawn at the sheet's DEFAULT
                # height, so that is what it is measured against.
                if not merged:
                    skipped.append(f"{ws.title}!{cell.coordinate}")
                    continue
                shipped = float(
                    ws.sheet_format.defaultRowHeight or DEFAULT_ROW_HEIGHT
                )
                # ...with one exemption, and only one: the default row height
                # is what Excel calibrates to ONE line of the DEFAULT font.
                # required_row_height returns the measured line BOX (16.0 pt at
                # font 11) which is a point over the 15.0 pt default row, and
                # reporting the dashboard's own one-line section headings as
                # clipped would be this gate's first false positive. A
                # multi-line merged cell, or one at a larger font, gets no such
                # pass — that is the case where lines 2..n genuinely never
                # display.
                lines = wrapped_line_count(
                    cell.value, chars_per_line(span, font_size)
                )
                if lines == 1 and font_size <= DEFAULT_FONT_SIZE:
                    checked += 1
                    continue
            shipped = float(shipped)
            checked += 1
            if needed > MAX_ROW_HEIGHT:
                findings.append(
                    f"{ws.title}!{cell.coordinate} needs {needed:.1f} pt, which "
                    f"is past Excel's {MAX_ROW_HEIGHT:.0f} pt row ceiling. No "
                    f"height can display it; split it across rows. "
                    f"({len(cell.value)} chars at {span:.0f} pt wide, "
                    f"font {font_size:g})"
                )
            elif shipped + 0.01 < needed:
                findings.append(
                    f"{ws.title}!{cell.coordinate} ships {shipped:.1f} pt and "
                    f"needs {needed:.1f} pt — the label is clipped and the "
                    f"reader sees only part of it. "
                    f"({len(cell.value)} chars at {span:.0f} pt wide, "
                    f"font {font_size:g}) {cell.value[:70]!r}"
                )
    return findings, checked, skipped


# ---------------------------------------------------------------------------
# THE THREE ANALYZER STATES, ALL BUILT.
#
# 1.3.0's fix set A12's height from the text and left the two BANNER rows at a
# hardcoded ``height = 28`` — in the same function, two rows above the cell it
# fixed. Nothing in the suite had ever built the workbook with a banner, so
# both were invisible. Measured on the shipped builder:
#
#     partial-unverified   A4 ships 28.0 pt, needs 75.0 pt (519 chars, font 10)
#     degraded             A4 ships 28.0 pt, needs 30.0 pt (168 chars, font 10)
#
# What the partial banner never displayed, past its first two lines:
#
#     "... so each is a LOWER BOUND ... Do not submit until all project
#      locations are verified."
#
# A gate that only ever sees the nominal fixture is a gate with one eye shut,
# so every geometry check below runs against all three states.
# ---------------------------------------------------------------------------

STATES = ("nominal", "partial_unverified", "degraded")


def _analysis_in_state(application_analysis, state: str):
    """Return the analysis re-stated in one of the three renderable states."""
    import copy
    analysis = copy.deepcopy(application_analysis)
    pr = analysis.pipeline_result
    if state == "partial_unverified":
        pr.eligibility_data_status = "ok"
        pr.unverified_project_ids = [p.project_id for p in list(pr.__dict__.get(
            "_projects", []))] or ["PIPE-0018", "PIPE-0019"]
    elif state == "degraded":
        pr.eligibility_data_status = "error"
        pr.eligibility_data_error = (
            "CDFI Fund eligibility table download failed: connection reset"
        )
    return analysis


@pytest.fixture(params=STATES)
def workbook(request, sample_application, application_analysis):
    """The workbook in each analyzer state — nominal, partial, degraded."""
    analysis = _analysis_in_state(application_analysis, request.param)
    wb = ExcelApplicationBuilder(sample_application, analysis).build()
    wb._nmtc_state = request.param
    return wb


def test_every_label_fits_its_row(workbook):
    """No label anywhere in the workbook may be taller than its own row."""
    findings = []
    checked = 0
    skipped = []
    for ws in workbook.worksheets:
        f, c, sk = check_sheet_geometry(ws)
        findings.extend(f)
        checked += c
        skipped.extend(sk)
    # VACUITY, HOLE 1: with no sheets, or no text cells, `findings` is empty
    # and the gate reports success on a workbook it never opened.
    assert workbook.worksheets, "the workbook has no sheets at all"
    assert checked > 0, (
        f"{workbook._nmtc_state}: not one text cell was measured. The gate "
        "just passed on a workbook it did not read, which is the vacuity it "
        "exists to refuse."
    )
    # VACUITY, HOLE 2: every height unset means every cell skipped, and the
    # gate passes silently. The builder sets an explicit height on every row it
    # writes — measured, zero skips — so a skip means that stopped being true
    # and the gate has quietly stopped covering those rows.
    assert not skipped, (
        f"{workbook._nmtc_state}: {len(skipped)} text cell(s) sit in unmerged "
        "rows with no explicit height, so this gate did not measure them. The "
        "builder sets a height on every row it writes; if that changed on "
        "purpose, re-derive this assertion rather than deleting it.\n  "
        + "\n  ".join(skipped[:20])
    )
    assert not findings, (
        f"{workbook._nmtc_state}: {len(findings)} label(s) do not fit their "
        f"rows, out of {checked} measured. This is 1.3.0 B1's class: a string "
        "that is correct in the source and clipped on the page. "
        "tests/rendered_baseline/excel.txt cannot see it — it records cell "
        "values, not geometry — so this gate is the only thing that can.\n\n"
        + "\n\n".join(findings)
    )


def test_the_banner_displays_its_whole_instruction(workbook):
    """B2's own cells, named, in the two states that render them.

    The partial banner ends "Do not submit until all project locations are
    verified." A CDE who never sees that sentence files anyway.
    """
    ws = workbook["Summary Dashboard"]
    banner = ws["A4"]
    if workbook._nmtc_state == "nominal":
        # ASSERTED, NOT SKIPPED. A skip here would be a third of this test
        # reporting success while checking nothing, in the file whose subject
        # is checks that cannot fail — and the negative is worth holding: a
        # banner rendered in the nominal state would be a false warning on
        # every clean run.
        assert banner.value is None, (
            "a banner rendered at A4 with nothing wrong: eligibility data "
            f"loaded and no project unverified, yet A4 reads {banner.value!r}"
        )
        return
    assert isinstance(banner.value, str) and banner.value.strip(), (
        f"{workbook._nmtc_state}: no banner rendered at A4 — the state that "
        "most needs a warning is the state with none"
    )
    findings, _, _ = check_sheet_geometry(ws)
    assert not findings, "\n".join(findings)


def test_the_distress_label_carries_its_whole_disclosure(workbook):
    """B1's own cell, named, so a regression reads as itself and not as drift."""
    ws = workbook["Summary Dashboard"]
    label = next(
        (c.value for c in ws["A"]
         if isinstance(c.value, str)
         and c.value.startswith("Deep/Severe Distress Concentration")),
        None,
    )
    assert label is not None, "the distress label is gone from the dashboard"
    assert "not of QLICIs" in label, (
        "the distress label no longer states its own denominator. That "
        "disclosure is the reason the row exists in this gate."
    )
    assert not check_sheet_geometry(ws)[0], (
        "the dashboard's distress label does not fit its row — see "
        "test_every_label_fits_its_row"
    )


def test_checker_catches_the_1_3_0_defect(workbook):
    """Rebuild B1 and prove the checker fails on it.

    A gate that has never been seen to fail is a gate nobody has evidence
    about. This restores exactly what shipped — a two-line label at
    ``height = 18`` — and asserts the checker reports it.
    """
    ws = workbook["Summary Dashboard"]
    target = next(
        c for c in ws["A"]
        if isinstance(c.value, str)
        and c.value.startswith("Deep/Severe Distress Concentration")
    )
    assert not check_sheet_geometry(ws)[0], "sheet must be clean before the defect"

    original = ws.row_dimensions[target.row].height
    try:
        ws.row_dimensions[target.row].height = 18  # the shipped 1.3.0 value
        findings = check_sheet_geometry(ws)[0]
        assert findings, (
            "the checker did not flag a two-line label at a one-line height — "
            "it cannot catch the defect it exists for"
        )
        assert any(f"A{target.row}" in f and "ships 18.0 pt" in f
                   for f in findings), findings
    finally:
        ws.row_dimensions[target.row].height = original

    assert not check_sheet_geometry(ws)[0], "the defect must not outlive this test"


def test_the_basis_note_is_reachable_by_name(workbook):
    """The dashboard's pointer must name something the workbook actually has.

    1.3.0 S4's pointer said "see the basis note below" and pointed fifteen rows
    down, past the readiness block and under the footer, at a row number
    nothing on screen names — and whether "below" was even visible depended on
    the reader's window, because the file stores no window geometry at all.
    """
    from nmtcapp.renderers._question_25 import Q25_BASIS_SHEET_NAME

    ws = workbook["Summary Dashboard"]
    label = next(
        c.value for c in ws["A"]
        if isinstance(c.value, str)
        and c.value.startswith("Deep/Severe Distress Concentration")
    )
    assert Q25_BASIS_SHEET_NAME in label, (
        "the distress label does not name the sheet its basis note is on"
    )
    assert Q25_BASIS_SHEET_NAME in workbook.sheetnames, (
        f"the label points at a sheet named {Q25_BASIS_SHEET_NAME!r} that this "
        "workbook does not contain"
    )
    assert workbook.sheetnames.index(Q25_BASIS_SHEET_NAME) == 1, (
        "the basis note sheet must sit immediately after the Summary "
        "Dashboard — a pointer is only as good as how far the reader looks"
    )


# ---------------------------------------------------------------------------
# THE THREE VACUITY HOLES, EACH CLOSED AND EACH PROVEN TO BE CLOSED
#
# Found by the confirmation pass on the 1.3.0 gate: it could report success on
# a workbook with no sheets, on a workbook where every height had been unset,
# and on a merged cell with no height — the last on a premise
# renderers/_sheet_geometry's own docstring contradicts in capitals. All three
# are the same shape as the defect the gate was built for: a check that cannot
# fail reads exactly like a check that passed.
# ---------------------------------------------------------------------------

def test_the_gate_refuses_a_workbook_with_no_sheets():
    """HOLE 1. Zero sheets means zero findings means, before this, a pass."""
    import openpyxl

    empty = openpyxl.Workbook()
    empty.remove(empty.active)
    empty._nmtc_state = "synthetic-empty"
    assert not empty.worksheets
    with pytest.raises(AssertionError, match="no sheets at all"):
        test_every_label_fits_its_row(empty)


def test_the_gate_refuses_a_workbook_whose_heights_were_all_unset(workbook):
    """HOLE 2. Unset every height and the old checker measured nothing at all.

    ``if not shipped: continue`` skipped the cell silently, so a builder that
    stopped setting heights would have turned the gate off rather than turned
    it red. The skip is now reported and asserted against.
    """
    for ws in workbook.worksheets:
        for row_dim in list(ws.row_dimensions.values()):
            row_dim.height = None

    with pytest.raises(AssertionError, match="no explicit height"):
        test_every_label_fits_its_row(workbook)


def test_the_gate_measures_a_merged_cell_that_was_never_given_a_height(workbook):
    """HOLE 3. Merged ranges do not autofit — the skip's premise was false.

    ``_sheet_geometry``'s docstring: "MERGED CELLS DO NOT AUTOFIT ... Excel's
    AutoFit is a no-op on a merged range, and every label on the dashboard is
    merged". The gate skipped exactly those cells on the opposite premise. This
    puts a multi-line label into a merged range with no height — the case where
    lines 2..n are simply never displayed — and asserts it is caught.
    """
    ws = workbook["Summary Dashboard"]
    row = ws.max_row + 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    from openpyxl.styles import Font
    cell = ws.cell(row=row, column=1, value="Merged and never autofitted. " * 12)
    cell.font = Font(size=10)
    assert ws.row_dimensions[row].height is None

    findings, checked, skipped = check_sheet_geometry(ws)
    assert not skipped, "a merged cell must never be skipped"
    assert any(f"A{row}" in f for f in findings), (
        "a 12-line label in a merged range at Excel's default row height was "
        "not reported. Every dashboard label is merged, so a gate that waves "
        "them through is a gate that covers none of them.\n" + "\n".join(findings)
    )
    assert checked > 0
