"""Tests for ExcelApplicationBuilder."""
import pytest

pytest.importorskip("openpyxl", reason="openpyxl not installed")

from nmtcapp.renderers.excel_builder import ExcelApplicationBuilder


class TestExcelBuilder:
    @pytest.fixture
    def builder(self, sample_application, application_analysis):
        return ExcelApplicationBuilder(sample_application, application_analysis)

    @pytest.fixture
    def wb(self, builder):
        return builder.build()

    # ---- Sheet count and names ----

    def test_has_nine_sheets(self, wb):
        # Eight since 1.3.0 B1: the Q25 basis note moved off the Summary
        # Dashboard onto its own sheet, which the dashboard's distress label
        # names. It was two rows at the bottom of the dashboard, pointed at by
        # "see the basis note below" — a direction whose truth depends on the
        # reader's window size, because the workbook stores no window geometry.
        #
        # NINE SINCE 1.5.0 B1: 'Round Provenance'. The workbook was the only
        # one of the four formats carrying no round provenance at all, while
        # still citing CY 2024-2025 in the present tense on the Q25 sheet.
        assert len(wb.sheetnames) == 9, wb.sheetnames

    def test_round_provenance_sheet_exists(self, wb):
        assert "Round Provenance" in wb.sheetnames

    def test_summary_dashboard_exists(self, wb):
        assert "Summary Dashboard" in wb.sheetnames

    def test_q25_basis_sheet_exists(self, wb):
        assert "Q25 Basis Note" in wb.sheetnames

    def test_pipeline_sheet_exists(self, wb):
        assert "Pipeline Detail" in wb.sheetnames

    def test_distress_sheet_exists(self, wb):
        assert "Distress Documentation" in wb.sheetnames

    def test_geographic_sheet_exists(self, wb):
        assert "Geographic Targeting" in wb.sheetnames

    def test_impact_sheet_exists(self, wb):
        assert "Impact Projections" in wb.sheetnames

    def test_investor_sheet_exists(self, wb):
        assert "Investor Commitments" in wb.sheetnames

    def test_track_record_sheet_exists(self, wb):
        assert "Track Record" in wb.sheetnames

    # ---- Summary Dashboard content ----

    def test_dashboard_has_cde_name(self, wb, sample_application):
        ws = wb["Summary Dashboard"]
        all_values = [ws.cell(row=r, column=c).value for r in range(1, 10) for c in range(1, 7)]
        texts = " ".join(str(v) for v in all_values if v)
        assert sample_application.cde.name in texts

    def test_dashboard_first_cell_has_title(self, wb):
        ws = wb["Summary Dashboard"]
        assert "NEW MARKETS" in str(ws["A1"].value).upper()

    def test_dashboard_has_metric_rows(self, wb):
        ws = wb["Summary Dashboard"]
        # Rows 6–14 should contain metrics
        non_empty = sum(
            1 for r in range(6, 15)
            if ws.cell(row=r, column=1).value
        )
        assert non_empty >= 5

    # ---- Pipeline sheet structure ----

    def test_pipeline_sheet_has_header_row(self, wb):
        ws = wb["Pipeline Detail"]
        # Row 3 is header; at least first cell has content
        assert ws.cell(row=3, column=1).value is not None

    def test_pipeline_sheet_has_data(self, wb, sample_application):
        ws = wb["Pipeline Detail"]
        # Row 4 onward should have project data
        assert ws.cell(row=4, column=1).value is not None

    def test_pipeline_sheet_auto_filter(self, wb):
        ws = wb["Pipeline Detail"]
        assert ws.auto_filter.ref is not None

    def test_pipeline_sheet_freeze_panes(self, wb):
        ws = wb["Pipeline Detail"]
        assert ws.freeze_panes is not None

    # ---- Geographic sheet ----

    def test_geographic_has_state_data(self, wb):
        ws = wb["Geographic Targeting"]
        # Header should mention "State" somewhere
        header_texts = " ".join(
            str(ws.cell(row=3, column=c).value or "") for c in range(1, 15)
        )
        assert "State" in header_texts or "state" in header_texts.lower()

    # ---- Save ----

    def test_save_creates_xlsx(self, builder, tmp_path):
        import os
        path = str(tmp_path / "application.xlsx")
        builder.save(path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 1000

    def test_save_creates_valid_xlsx(self, builder, tmp_path):
        import os
        from openpyxl import load_workbook
        path = str(tmp_path / "application.xlsx")
        builder.save(path)
        wb2 = load_workbook(path)
        assert len(wb2.sheetnames) == 9, wb2.sheetnames

    def test_save_creates_parent_directory(self, builder, tmp_path):
        import os
        path = str(tmp_path / "output" / "application.xlsx")
        builder.save(path)
        assert os.path.exists(path)


class TestExcelBuilderImportError:
    def test_raises_if_openpyxl_missing(self, sample_application, application_analysis,
                                         monkeypatch):
        import nmtcapp.renderers.excel_builder as mod
        monkeypatch.setattr(mod, "_OPENPYXL_AVAILABLE", False)
        with pytest.raises(ImportError, match="openpyxl"):
            ExcelApplicationBuilder(sample_application, application_analysis)
