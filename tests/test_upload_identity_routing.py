"""T1 (1.6.0): the CDE's own identity must reach ``CDEProfile``.

ONE ROOT CAUSE, FOUR SYMPTOMS.

``streamlit_app.utils._IDENTITY_KEYS`` strips ``cde_name``, ``mission``,
``certification_date``, ``target_markets``, ``website``, ``ein``,
``headquarters_state``, ``organization_type`` and ``cde_id`` out of the parsed
CDE Profile sheet. THE STRIP IS CORRECT AND STAYS: 1.1.5 proved what happens
when identity merges into the SCORING BAG -- a CDE that filled only the
Pipeline sheet inherited the fictional Riverbend CDE's
``has_prior_reporting_issues: False``, and ``sections/base._compliance_statement``
rendered a clean-compliance claim into a federal filing.

What was missing is the other half. 1.5.7 routed the ROUND and the ALLOCATION
past the strip via ``UploadedCDEProfile``, to ``Application``. It routed
nothing else. So ``get_or_create_app``'s neutral profile still won, and a CDE
that typed its name into the template's own "CDE Name" cell got a federal
filing draft that said ``(your CDE)`` on every page and in the filename.

MEASURED AT 9a2d584, on a workbook filled exactly as the shipped blank
template instructs:

    (your CDE)                 x8      filename: user-upload_application.md
    completeness component     0.0/100
    completeness_check issues  certification_date, mission, target_markets,
                               contact, governance

Three of those five the CDE SUPPLIED and the strip removed. The completeness
check was reporting the truth; the truth was the defect.

IDENTITY REACHING ``CDEProfile`` IS NOT IDENTITY MERGING INTO THE SCORING BAG.
Identity travels on its own field of ``UploadedCDEProfile`` and lands on
``CDEProfile``'s own named attributes. ``scoring_attrs`` is untouched, so the
1.1.5 defence is unchanged -- which
``test_the_scoring_bag_still_carries_no_identity`` below asserts directly.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest

_REPO_ROOT = Path(__file__).resolve().parent.parent
_STREAMLIT_APP = _REPO_ROOT / "streamlit_app"
for _p in (str(_REPO_ROOT), str(_STREAMLIT_APP)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import streamlit as st  # noqa: E402

from nmtcapp.core.upload_handler import load_uploaded_pipeline  # noqa: E402
from nmtcapp.core.sample_identity import SampleDataError  # noqa: E402

from tests.conftest import templates_dir  # noqa: E402

from utils import (  # noqa: E402
    get_or_create_app,
    read_uploaded_cde_profile,
)

#: What the CDE types. None of it is the shipped sample's, and none of it is
#: the neutral profile's placeholder, so a surface rendering either is
#: rendering something the uploader did not supply.
TYPED = {
    "CDE Name": "Cardinal Ridge Community Capital, LLC",
    "CDE ID": "CDE-2020-0431",
    "EIN": "84-3319027",
    "Certification Date": "2020-09-14",
    "Mission Statement": (
        "Deploy New Markets Tax Credit capital into persistently distressed "
        "rural and small-metro communities across the Carolinas and "
        "Appalachia."
    ),
    "Website": "https://cardinalridgecapital.org",
    "Target Markets (states, comma-sep)": "North Carolina, South Carolina, Tennessee",
}


def _workbook() -> bytes:
    wb = openpyxl.load_workbook(Path(templates_dir()) / "pipeline_sample.xlsx")
    ws = wb["CDE Profile"]
    headers = {str(ws.cell(3, c).value).strip(): c
               for c in range(1, ws.max_column + 1) if ws.cell(3, c).value}
    for label, value in TYPED.items():
        assert label in headers, (
            f"the shipped template no longer has a {label!r} column; this "
            "gate has lost its subject"
        )
        ws.cell(row=4, column=headers[label]).value = value
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture(autouse=True)
def _fresh_session():
    st.session_state.clear()
    yield
    st.session_state.clear()


@pytest.fixture(scope="module")
def uploaded() -> tuple:
    return load_uploaded_pipeline(_workbook(), "cardinal_ridge.xlsx")


def _app(uploaded):
    pipeline, cde_extra = uploaded
    parsed = read_uploaded_cde_profile(cde_extra, is_demo=False)
    return get_or_create_app(pipeline=pipeline, is_demo=False, cde_extra=parsed)


# ---------------------------------------------------------------------------
# Symptom 1 -- the placeholder name
# ---------------------------------------------------------------------------

class TestTheCDEsOwnIdentityReachesTheProfile:
    def test_the_name_the_cde_typed_is_the_profile_name(self, uploaded):
        assert _app(uploaded).cde.name == TYPED["CDE Name"]

    def test_the_cde_id_the_cde_typed_is_the_profile_id(self, uploaded):
        assert _app(uploaded).cde.cde_id == TYPED["CDE ID"]

    def test_the_certification_date_reaches_the_profile(self, uploaded):
        assert _app(uploaded).cde.certification_date == TYPED["Certification Date"]

    def test_the_mission_reaches_the_profile(self, uploaded):
        assert _app(uploaded).cde.mission == TYPED["Mission Statement"]

    def test_the_website_reaches_the_profile(self, uploaded):
        assert _app(uploaded).cde.website == TYPED["Website"]

    def test_target_markets_reach_the_profile_as_a_list(self, uploaded):
        assert _app(uploaded).cde.target_markets == [
            "North Carolina", "South Carolina", "Tennessee",
        ]


# ---------------------------------------------------------------------------
# T1b -- the strip's purpose survives
# ---------------------------------------------------------------------------

class TestTheStripsPurposeSurvives:
    def test_the_scoring_bag_still_carries_no_identity(self, uploaded):
        """The 1.1.5 defence, unchanged: identity may not reach ``extra``."""
        from utils import _IDENTITY_KEYS

        extra = _app(uploaded).cde.extra
        leaked = sorted(set(extra) & set(_IDENTITY_KEYS))
        assert not leaked, (
            "identity keys reached the SCORING BAG. Identity may land on "
            "CDEProfile's own attributes; it may never merge into "
            f"CDEProfile.extra. Leaked: {leaked}"
        )

    def test_an_upload_that_names_no_cde_still_gets_the_neutral_profile(self):
        """Routing identity may not invent one for an upload that states none."""
        from nmtcapp.core.pipeline import Pipeline

        app = get_or_create_app(pipeline=Pipeline.sample(n=4), is_demo=False,
                                cde_extra=None)
        assert app.cde.name == "(your CDE)"
        assert app.cde.cde_id == "user-upload"

    def test_a_sheet_carrying_the_sample_identity_is_still_refused(self, uploaded):
        pipeline, _ = uploaded
        with pytest.raises(SampleDataError):
            read_uploaded_cde_profile(
                {"cde_name": "Riverbend Community Capital CDE, LLC",
                 "cde_id": "CDE-2018-0117"},
                is_demo=False,
            )


class TestTheIsDemoSeam:
    """Seam 1 of the two the 1.5.7 audit named on this code.

    ``is_demo`` is supplied TWICE -- once to ``read_uploaded_cde_profile``,
    which decides whether ``assert_not_sample_identity`` runs, and once to
    ``get_or_create_app``, which decides whether the sample profile is used.
    Nothing checked the two agreed, so the guard was bypassable::

        parsed = read_uploaded_cde_profile(sample_sheet, is_demo=True)   # skips
        get_or_create_app(pipeline=p, is_demo=False, cde_extra=parsed)   # merges

    Latent -- page 1 passes consistent values -- and closed structurally
    rather than left to depend on that.
    """

    def test_a_profile_read_as_demo_cannot_be_used_as_a_real_upload(self):
        from nmtcapp.core.pipeline import Pipeline

        parsed = read_uploaded_cde_profile(
            {"cde_name": "Riverbend Community Capital CDE, LLC",
             "cde_id": "CDE-2018-0117",
             "dbc_focus_years": 4},
            is_demo=True,
        )
        with pytest.raises(AssertionError):
            get_or_create_app(pipeline=Pipeline.sample(n=4), is_demo=False,
                              cde_extra=parsed)

    def test_a_profile_read_as_a_real_upload_cannot_be_used_as_the_demo(self, uploaded):
        pipeline, cde_extra = uploaded
        parsed = read_uploaded_cde_profile(cde_extra, is_demo=False)
        with pytest.raises(AssertionError):
            get_or_create_app(pipeline=pipeline, is_demo=True, cde_extra=parsed)

    def test_consistent_values_are_accepted(self, uploaded):
        assert _app(uploaded).cde.name == TYPED["CDE Name"]


class TestDualShapeAcceptanceIsGone:
    """Seam 2. RULED IN 1.6.0: the raw-dict shape no longer resolves.

    ``get_or_create_app`` accepted ``dict | UploadedCDEProfile``, and the
    1.5.7 audit proved the type was OPTIONAL ARMOUR -- the page-driving gate,
    not the type, is what closed the class. That was defensible while the
    function could redo the read itself. After T1 it cannot: redoing the read
    means running ``_scoring_attrs_only``, whose sample-identity guard is
    gated on an ``is_demo`` the caller supplies SEPARATELY -- which is seam 1,
    the very thing being closed above. A dict caller would also silently lose
    the CDE's identity, which is the 1.5.6 defect re-armed one field wider.

    ``streamlit_app`` is not shipped (``pyproject`` packages ``nmtcapp*``
    only), so this is an app-internal signature, not a public API break.
    """

    def test_a_raw_dict_is_refused(self):
        from nmtcapp.core.pipeline import Pipeline

        with pytest.raises(TypeError):
            get_or_create_app(pipeline=Pipeline.sample(n=4), is_demo=False,
                              cde_extra={"dbc_focus_years": 4})
