"""Round-trip test: v1.1 xlsx template → fill 3 rows → save as bytes → upload → Pipeline.

Tests the primary user journey:
  1. User downloads pipeline_template.xlsx
  2. Fills in project data
  3. Saves and uploads the xlsx back to the Pipeline Analyzer
  4. Gets a valid Pipeline object

Uses load_uploaded_pipeline() from nmtcapp.core.upload_handler so no Streamlit
runtime is required.
"""
from __future__ import annotations

import io
import os
import sys

import openpyxl
import pytest

_REPO_ROOT = os.path.dirname(os.path.dirname(__file__))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from nmtcapp.core.upload_handler import load_uploaded_pipeline

_TEMPLATE_PATH = os.path.join(_REPO_ROOT, "templates", "pipeline_template.xlsx")

# ---------------------------------------------------------------------------
# 3 test rows — match the Pipeline sheet's column order (row 3 headers):
# Project ID | Project Name | QALICB Name | Address | City | State |
# Sector | Project Type | QEI ($M) | Total Cost ($M) | Jobs Created |
# Jobs Retained | Affordable Units | Commercial Sq Ft | Distress Level |
# Native Area (Y/N) | High Migration Rural (Y/N) | US Territory (Y/N) |
# Persistent Poverty (Y/N) | Below-Market Rate (Y/N) | Unrelated Entity (Y/N) |
# Opportunity Zone (Y/N) | Urban/Rural | Census Tract (11-digit) |
# Closing Target Date | Minority Owned (Y/N) | Women Owned (Y/N) | Notes
# ---------------------------------------------------------------------------
_TEST_ROWS = [
    (
        "PRJ-T001", "Roundtrip Health Center", "RT HC QALICB LLC",
        "100 Main St", "Chicago", "IL",
        "healthcare", "real_estate",
        5.0, 8.0, 25, 10, None, None, None,
        "N", "N", "N", "N", "Y", "Y", "N",
        "urban", None, "2025-12-31", "N", "N", None,
    ),
    (
        "PRJ-T002", "Roundtrip Charter School", "RT School QALICB LLC",
        "200 Oak Ave", "Houston", "TX",
        "education", "real_estate",
        4.0, 6.0, 30, 5, None, None, None,
        "N", "N", "N", "Y", "Y", "Y", "N",
        "urban", None, "2026-03-31", "N", "N", None,
    ),
    (
        "PRJ-T003", "Roundtrip Factory", "RT MFG QALICB LLC",
        "300 Industrial Blvd", "Detroit", "MI",
        "small_business", "operating_business",
        3.0, 5.0, 50, 20, None, None, None,
        "N", "N", "N", "N", "Y", "N", "N",
        "urban", None, "2026-06-30", "N", "N", None,
    ),
]


def _build_xlsx_with_3_rows() -> bytes:
    """Load the template, clear sample data rows, write 3 test rows, return bytes."""
    wb = openpyxl.load_workbook(_TEMPLATE_PATH)
    ws = wb["Pipeline"]

    # Clear all rows from row 4 onward (sample data the template ships with)
    for r in range(4, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

    # Write test rows starting at row 4
    for row_offset, row_data in enumerate(_TEST_ROWS):
        r = 4 + row_offset
        for col_offset, val in enumerate(row_data):
            ws.cell(r, 1 + col_offset).value = val

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def xlsx_bytes_3rows():
    if not os.path.exists(_TEMPLATE_PATH):
        pytest.skip(f"Template not found: {_TEMPLATE_PATH}")
    return _build_xlsx_with_3_rows()


class TestTemplateRoundTrip:
    """Primary user journey: download template → fill → upload → Pipeline."""

    def test_upload_produces_pipeline_with_3_projects(self, xlsx_bytes_3rows):
        pipeline, _ = load_uploaded_pipeline(xlsx_bytes_3rows, "my_pipeline.xlsx")
        assert len(pipeline) == 3, (
            f"Expected 3 projects from uploaded xlsx, got {len(pipeline)}. "
            "Check that the Pipeline sheet header is read at row 3 (header=2) "
            "and that the column mapping in _XLSX_PIPELINE_COL_MAP is correct."
        )

    def test_qei_and_cost_converted_from_millions(self, xlsx_bytes_3rows):
        """QEI ($M) and Total Cost ($M) in the template must be multiplied × 1_000_000."""
        pipeline, _ = load_uploaded_pipeline(xlsx_bytes_3rows, "my_pipeline.xlsx")
        projects = list(pipeline)

        # PRJ-T001: QEI ($M)=5.0 → qei_request=5_000_000; Total Cost ($M)=8.0 → 8_000_000
        assert projects[0].qei_request == pytest.approx(5_000_000.0), (
            f"qei_request should be 5_000_000 (5.0M × 1e6), got {projects[0].qei_request}"
        )
        assert projects[0].total_project_cost == pytest.approx(8_000_000.0), (
            f"total_project_cost should be 8_000_000, got {projects[0].total_project_cost}"
        )
        assert projects[0].expected_jobs_created == 25
        assert projects[0].state == "IL"
        assert projects[0].project_id == "PRJ-T001"

        # PRJ-T003: QEI=3.0M, Cost=5.0M
        assert projects[2].qei_request == pytest.approx(3_000_000.0)
        assert projects[2].total_project_cost == pytest.approx(5_000_000.0)
        assert projects[2].expected_jobs_created == 50

    def test_cde_profile_sheet_parses_without_error(self):
        """The template's CDE Profile sheet should parse to a non-empty dict."""
        if not os.path.exists(_TEMPLATE_PATH):
            pytest.skip(f"Template not found: {_TEMPLATE_PATH}")
        with open(_TEMPLATE_PATH, "rb") as f:
            xlsx_bytes = f.read()
        _pipeline, cde_extra = load_uploaded_pipeline(xlsx_bytes, "pipeline_template.xlsx")
        assert cde_extra is not None and len(cde_extra) >= 1, (
            "CDE Profile sheet should parse to a dict with at least one field. "
            f"Got: {cde_extra!r}"
        )
