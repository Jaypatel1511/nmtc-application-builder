"""Template / scoring-engine field alignment tests (v1.1).

Two tests:
  test_template_has_all_methodology_fields
    Load nmtcapp/templates/pipeline_template.xlsx and verify that every field the
    scoring engine needs is present in the correct sheet.

  test_scoring_engine_inputs_match_template
    Cross-reference the canonical list of scoring-engine input keys against
    both the template and the WinProbabilityModel implementation to confirm
    they stay in sync.
"""
from __future__ import annotations

import os
import sys

_REPO_ROOT = os.path.dirname(os.path.dirname(__file__))

from tests.conftest import templates_dir as _templates_dir
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

import openpyxl
import pytest

# ---------------------------------------------------------------------------
# Canonical scoring-engine input sets
# ---------------------------------------------------------------------------

# Fields the scoring engine reads from the Pipeline sheet (column names in
# the v1.1 template, row 3).  These map to PipelineProject attributes and are
# consumed by distress_analysis / geographic_analysis.
PIPELINE_REQUIRED_COLUMNS = {
    "Project ID",
    "Project Name",
    "QALICB Name",
    "State",
    "City",
    "Sector",
    "Project Type",
    "QEI ($M)",
    "Total Cost ($M)",
    "Jobs Created",
    "Distress Level",
}

PIPELINE_SCORING_FLAGS = {
    # Per-project flags that drive CDE-level scoring pcts
    "Native Area (Y/N)",
    "High Migration Rural (Y/N)",
    "US Territory (Y/N)",
    "Persistent Poverty (Y/N)",
    "Below-Market Rate (Y/N)",
    "Unrelated Entity (Y/N)",
}

# CDE Profile fields that feed WinProbabilityModel.score(cde_attributes=...)
CDE_SCORING_FIELDS = {
    # Business Strategy
    "Below-Market Rate Pct (0–1)*",
    "Flexible Product Indicia Count",
    "Pipeline % Identified (0–1)",
    "Own Capital at Risk (Y/N)",
    "Prior Award Count",
    "Years in Operation",
    "Track Record Pipeline Align (0–1)",
    "Track Record Deployment Pct (0–1)",
    # Community Outcomes
    "Persistent Poverty Pct (0–1)*",
    "US Territories Pct (0–1)*",
    "Quantified Outcomes (Y/N)",
    "Third-Party Validation (Y/N)",
    "LIC Board Representation (0–1)",
    "Community Engagement Track Record (Y/N)",
    # Priority Points
    "DBC Focus Years",
    "DBC Dollar Volume Pct (0–1)*",
    "Unrelated Entities Pct (0–1)*",
    # Phase 2
    "Favorable Fee Structure (Y/N/Unknown)",
    "Prior Reporting Issues (Y/N)",
}

# Canonical attrs keys consumed by WinProbabilityModel.score()
SCORING_ENGINE_ATTRS_KEYS = {
    "products_below_market_pct",
    "products_flexible_indicia_count",
    "pipeline_pct_identified",
    "has_own_capital_at_risk",
    "prior_award_count",
    "years_in_operation",
    "track_record_pipeline_alignment_pct",
    "track_record_deployment_pct",
    "pct_persistent_poverty",
    "pct_us_territories",
    "has_quantified_outcomes",
    "has_third_party_validation",
    "lic_board_representation_pct",
    "has_community_engagement_track_record",
    "dbc_focus_years",
    "dbc_dollar_volume_pct",
    "unrelated_entities_pct",
    "has_favorable_fee_structure",
    "has_prior_reporting_issues",
}

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

_TEMPLATE_PATH = os.path.join(_templates_dir(), "pipeline_template.xlsx")


@pytest.fixture(scope="module")
def template_wb():
    assert os.path.exists(_TEMPLATE_PATH), (
        f"Template not found at {_TEMPLATE_PATH} — run the build script first."
    )
    return openpyxl.load_workbook(_TEMPLATE_PATH, data_only=True)


@pytest.fixture(scope="module")
def pipeline_headers(template_wb):
    """Row 3 headers from the Pipeline sheet (the column-label row)."""
    ws = template_wb["Pipeline"]
    return {ws.cell(3, c).value for c in range(1, ws.max_column + 1) if ws.cell(3, c).value}


@pytest.fixture(scope="module")
def cde_headers(template_wb):
    """Row 3 headers from the CDE Profile sheet."""
    ws = template_wb["CDE Profile"]
    return {ws.cell(3, c).value for c in range(1, ws.max_column + 1) if ws.cell(3, c).value}


# ---------------------------------------------------------------------------
# Test 1: template_has_all_methodology_fields
# ---------------------------------------------------------------------------

class TestTemplateHasAllMethodologyFields:
    """Every field the scoring engine needs must be present in the template."""

    def test_pipeline_sheet_exists(self, template_wb):
        assert "Pipeline" in template_wb.sheetnames, \
            "Pipeline sheet is missing from the template."

    def test_cde_profile_sheet_exists(self, template_wb):
        assert "CDE Profile" in template_wb.sheetnames, \
            "CDE Profile sheet is missing from the template."

    def test_instructions_sheet_exists(self, template_wb):
        assert "Instructions" in template_wb.sheetnames, \
            "Instructions sheet is missing."

    def test_valid_values_sheet_exists(self, template_wb):
        assert "Valid Values" in template_wb.sheetnames, \
            "Valid Values sheet is missing."

    def test_pipeline_required_columns_present(self, pipeline_headers):
        missing = PIPELINE_REQUIRED_COLUMNS - pipeline_headers
        assert not missing, (
            f"Pipeline sheet is missing required columns: {sorted(missing)}"
        )

    def test_pipeline_scoring_flags_present(self, pipeline_headers):
        missing = PIPELINE_SCORING_FLAGS - pipeline_headers
        assert not missing, (
            f"Pipeline sheet is missing v1.1 scoring flags: {sorted(missing)}\n"
            "These flags drive CDE-level pcts (native_area→pct_native_area, etc.)."
        )

    def test_cde_profile_scoring_fields_present(self, cde_headers):
        missing = CDE_SCORING_FIELDS - cde_headers
        assert not missing, (
            f"CDE Profile sheet is missing scoring fields: {sorted(missing)}"
        )

    def test_pipeline_has_qalicb_name(self, pipeline_headers):
        assert "QALICB Name" in pipeline_headers, \
            "Pipeline sheet must have 'QALICB Name' column (required by PipelineProject)."

    def test_pipeline_has_closing_date(self, pipeline_headers):
        assert "Closing Target Date" in pipeline_headers, \
            "Pipeline sheet must have 'Closing Target Date' column."

    def test_cde_profile_is_blank(self, template_wb):
        """The CDE Profile data row must be EMPTY. Inverted in 1.2.0.

        This test used to assert the opposite — "row 4 must be non-empty" —
        and so encoded the defect. The template shipped pre-filled with
        Riverbend Community Capital CDE, LLC, CDE-2018-0117, EIN 82-1234567
        and eighteen scoring attributes, and Streamlit served it as "Download
        blank template (Excel)". A CDE that filled in only the Pipeline sheet
        uploaded Riverbend's identity, and Riverbend's
        has_prior_reporting_issues=False rendered into a federal filing as
        "Per this CDE's own profile declaration, no prior NMTC reporting
        issues have been recorded."

        Populated example data lives in pipeline_sample.xlsx.
        """
        ws = template_wb["CDE Profile"]
        filled = [ws.cell(4, c).value for c in range(1, ws.max_column + 1)]
        non_empty = [v for v in filled if v is not None]
        assert not non_empty, (
            f"pipeline_template.xlsx ships {len(non_empty)} pre-filled CDE "
            f"Profile cell(s): {non_empty[:5]}. A file named *_template.* must "
            "be blank in every field a CDE is expected to supply."
        )

    def test_pipeline_is_blank(self, template_wb):
        """The Pipeline sheet must ship no example rows, for the same reason."""
        ws = template_wb["Pipeline"]
        filled = []
        for r in range(4, ws.max_row + 1):
            filled += [v for v in (ws.cell(r, c).value
                                   for c in range(1, ws.max_column + 1))
                       if v is not None]
        assert not filled, (
            f"pipeline_template.xlsx ships {len(filled)} pre-filled project "
            f"cell(s): {filled[:5]}. Example rows belong in pipeline_sample.xlsx."
        )

    def test_sample_workbook_exists_and_is_labelled(self):
        """The populated workbook survives, but says on its face that it is fake."""
        import openpyxl
        path = os.path.join(_templates_dir(), "pipeline_sample.xlsx")
        assert os.path.exists(path), (
            "pipeline_sample.xlsx is missing — blanking the template must not "
            "delete the worked example, only separate it."
        )
        wb = openpyxl.load_workbook(path)
        for sheet in ("CDE Profile", "Pipeline"):
            banner = str(wb[sheet]["A1"].value or "")
            assert "SAMPLE" in banner.upper() and "FICTIONAL" in banner.upper(), (
                f"{sheet}!A1 of pipeline_sample.xlsx does not announce itself as "
                f"fictional sample data. Got: {banner!r}"
            )
        assert [c.value for c in wb["CDE Profile"][4] if c.value is not None], (
            "pipeline_sample.xlsx should still carry its worked example"
        )


# ---------------------------------------------------------------------------
# Test 2: scoring_engine_inputs_match_template
# ---------------------------------------------------------------------------


# Map: scoring-engine attr key → CDE Profile display label (v1.1 template row 3)
_ATTR_KEY_TO_LABEL: dict[str, str] = {
    "products_below_market_pct":            "Below-Market Rate Pct (0–1)*",
    "products_flexible_indicia_count":      "Flexible Product Indicia Count",
    "pipeline_pct_identified":              "Pipeline % Identified (0–1)",
    "has_own_capital_at_risk":              "Own Capital at Risk (Y/N)",
    "prior_award_count":                    "Prior Award Count",
    "years_in_operation":                   "Years in Operation",
    "track_record_pipeline_alignment_pct":  "Track Record Pipeline Align (0–1)",
    "track_record_deployment_pct":          "Track Record Deployment Pct (0–1)",
    "pct_persistent_poverty":               "Persistent Poverty Pct (0–1)*",
    "pct_us_territories":                   "US Territories Pct (0–1)*",
    "has_quantified_outcomes":              "Quantified Outcomes (Y/N)",
    "has_third_party_validation":           "Third-Party Validation (Y/N)",
    "lic_board_representation_pct":         "LIC Board Representation (0–1)",
    "has_community_engagement_track_record": "Community Engagement Track Record (Y/N)",
    "dbc_focus_years":                      "DBC Focus Years",
    "dbc_dollar_volume_pct":                "DBC Dollar Volume Pct (0–1)*",
    "unrelated_entities_pct":              "Unrelated Entities Pct (0–1)*",
    "has_favorable_fee_structure":          "Favorable Fee Structure (Y/N/Unknown)",
    "has_prior_reporting_issues":           "Prior Reporting Issues (Y/N)",
}

# These attrs can be auto-computed from per-project Pipeline flags, so a CDE
# Profile column is optional (the scorer falls back to the pipeline value).
_PIPELINE_DERIVABLE = {
    "products_below_market_pct",
    "pct_us_territories",
    "pct_persistent_poverty",
    "unrelated_entities_pct",
    "dbc_dollar_volume_pct",
}


class TestScoringEngineInputsMatchTemplate:
    """The canonical attrs keys consumed by the scoring engine must all have
    corresponding columns in the CDE Profile sheet or Pipeline sheet."""

    def test_all_scoring_attrs_covered_by_template(self, cde_headers, pipeline_headers):
        """Every key in SCORING_ENGINE_ATTRS_KEYS must appear somewhere in the template."""
        missing_coverage = []
        for key in SCORING_ENGINE_ATTRS_KEYS:
            label = _ATTR_KEY_TO_LABEL.get(key)
            in_cde = label in cde_headers if label else False
            if not in_cde and key not in _PIPELINE_DERIVABLE:
                missing_coverage.append(key)
        assert not missing_coverage, (
            f"These scoring-engine attrs have no coverage in the template: "
            f"{sorted(missing_coverage)}"
        )

    def test_win_probability_model_reads_canonical_keys(self):
        """WinProbabilityModel must call attrs.get() for each canonical key."""
        import inspect
        from nmtcapp.intelligence.win_probability import WinProbabilityModel
        source = inspect.getsource(WinProbabilityModel)
        missing = [k for k in SCORING_ENGINE_ATTRS_KEYS if f'"{k}"' not in source]
        assert not missing, (
            f"WinProbabilityModel does not read these canonical attrs keys: "
            f"{sorted(missing)}\n"
            "Either add them to the scorer or remove them from SCORING_ENGINE_ATTRS_KEYS."
        )

    def test_distress_analysis_surfaces_pipeline_derived_keys(self):
        """distress_analysis must return the pipeline-derived pct fields."""
        from nmtcapp.core.pipeline import Pipeline
        from nmtcapp.intelligence.distress_analysis import analyze_distress_concentration
        result = analyze_distress_concentration(Pipeline.sample(n=5))
        pipeline_derived_keys = {
            "pct_us_territories",
            "pct_persistent_poverty",
            "pct_below_market_rate",
            "pct_unrelated_entity",
            "pct_native_area",
            "pct_high_migration_rural",
        }
        missing = pipeline_derived_keys - set(result.keys())
        assert not missing, (
            f"analyze_distress_concentration() is missing pipeline-derived keys: "
            f"{sorted(missing)}"
        )

    def test_scoring_engine_uses_pipeline_fallback_for_below_market(self):
        """When cde_attributes omits products_below_market_pct, the scorer
        falls back to pct_below_market_rate from the pipeline result."""
        from nmtcapp.core.pipeline import Pipeline
        from nmtcapp.intelligence.win_probability import WinProbabilityModel
        from nmtcapp.intelligence.pipeline_analyzer import PipelineAnalyzer

        pipeline = Pipeline.sample(n=5)
        pr = PipelineAnalyzer().analyze(pipeline)

        model = WinProbabilityModel()
        # Score with no attrs (should fall back to pipeline pct_below_market_rate)
        score_no_attrs = model.score(pr, 30_000_000, cde_attributes={})
        # Score with explicit below_market_pct=0
        score_zero_attr = model.score(
            pr, 30_000_000, cde_attributes={"products_below_market_pct": 0.0}
        )
        # Score with explicit below_market_pct=1.0 (full credit)
        score_full_attr = model.score(
            pr, 30_000_000, cde_attributes={"products_below_market_pct": 1.0}
        )

        pf_no_attrs = score_no_attrs.business_strategy["product_flexibility"]
        pf_zero = score_zero_attr.business_strategy["product_flexibility"]
        pf_full = score_full_attr.business_strategy["product_flexibility"]

        # Full credit (explicit 100%) must be higher than zero (explicit 0%)
        assert pf_full > pf_zero, (
            f"Explicit pct=1.0 should outscore explicit pct=0 ({pf_full} vs {pf_zero})"
        )
        # When no attrs provided, score should reflect pipeline flags (sample has below_market=True)
        assert pf_no_attrs >= pf_zero, (
            f"Pipeline fallback score ({pf_no_attrs}) should be >= zero-attr score ({pf_zero}) "
            "since sample pipeline projects have is_below_market_rate=True"
        )

    def test_scoring_engine_uses_pipeline_fallback_for_unrelated_entities(self):
        """When cde_attributes omits unrelated_entities_pct, the scorer
        falls back to pct_unrelated_entity from the pipeline result."""
        from nmtcapp.core.pipeline import Pipeline
        from nmtcapp.intelligence.win_probability import WinProbabilityModel
        from nmtcapp.intelligence.pipeline_analyzer import PipelineAnalyzer

        pipeline = Pipeline.sample(n=5)
        pr = PipelineAnalyzer().analyze(pipeline)
        model = WinProbabilityModel()

        score_no_attr = model.score(pr, 30_000_000, cde_attributes={})
        score_zero    = model.score(pr, 30_000_000, cde_attributes={"unrelated_entities_pct": 0.0})
        score_full    = model.score(pr, 30_000_000, cde_attributes={"unrelated_entities_pct": 1.0})

        ue_no_attr = score_no_attr.priority_points["unrelated_entities"]
        ue_zero    = score_zero.priority_points["unrelated_entities"]
        ue_full    = score_full.priority_points["unrelated_entities"]

        assert ue_full > ue_zero, (
            f"Explicit pct=1.0 unrelated should outscore 0 ({ue_full} vs {ue_zero})"
        )
        assert ue_no_attr >= ue_zero, (
            f"Pipeline fallback ({ue_no_attr}) should be >= zero ({ue_zero}) for sample pipeline"
        )
