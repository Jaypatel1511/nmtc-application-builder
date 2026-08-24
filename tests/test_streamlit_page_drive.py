"""RENDERED-OUTPUT GATE 3 of 3: does the PAGE, driven end to end, say what the
CDE typed?

WHY THIS FILE EXISTS (1.5.7 T2) -- AND WHY IT IS A NEW SHAPE

The eighth entry on this package's gates-that-cannot-fail ledger is not a dead
assertion and not a detector wired to its own subject. It is a CORRECT unit
with a CORRECT test, called in an order that made both irrelevant.

``streamlit_app/utils.get_or_create_app`` reads the round and the requested
allocation off the parsed CDE Profile sheet BEFORE it strips identity keys out
of that dict, and its comment says so: "Read before ``_scoring_attrs_only()``
below rebinds cde_extra without them." That is true, and it defends against
its OWN strip. ``tests/test_streamlit_upload_profile.py`` has EIGHTEEN tests
over exactly this behaviour, including
``test_an_upload_that_states_its_round_has_it_honoured`` and
``test_an_upload_that_states_its_allocation_has_it_honoured``. They pass. They
have always passed. They are correct.

Every one of them hands the function a dict that STILL CONTAINS THE KEYS,
because the test author wrote the dict by hand::

    get_or_create_app(..., cde_extra={"application_round": "CY 2027", ...})

``streamlit_app/pages/1_Pipeline_Analyzer.py`` called ``_scoring_attrs_only``
ONE LINE EARLIER and REBOUND the name, so what it handed over was a dict
already emptied of both keys. The function read it, found nothing, and fell
back to ``_UNSTATED_ALLOCATION_PLACEHOLDER``. A CDE that filled in the
template's own "Application Round" and "Requested Allocation ($M)" cells was
told on screen that it had supplied NEITHER -- and \\$65,000,000, a figure it
had never typed, was then used in the validators' own arithmetic:

    Application round:    not specified - CDE to state
    Requested allocation: not supplied [CDE TO COMPLETE]
    Total pipeline QEI (30,200,000) is below the requested allocation
    (65,000,000) - pipeline undersized

THE TESTS CONSTRUCT THE INPUT THE FUNCTION WANTS. THE PAGE CONSTRUCTS THE
INPUT THE FUNCTION GETS. No unit test over ``utils`` can see the difference,
because the difference IS the caller. Only something that drives the page can,
which is what this file does.

  !!  SCOPE LIMIT -- READ THIS BEFORE TRUSTING A GREEN FROM THIS FILE  !!

  WHAT IT RUNS. Streamlit's own ``AppTest`` harness, in-process, against the
  REAL page files under ``streamlit_app/pages/``. The page's module-level
  code, its radio, its button, ``load_uploaded_pipeline``,
  ``read_uploaded_cde_profile``, ``get_or_create_app``, ``Application.analyze``
  and every ``st.*`` call all execute. The assertions below read the text the
  page actually emitted, not the source that produced it.

  THE ONE SEAM. ``st.file_uploader``'s RETURN VALUE is stubbed with a
  ``BytesIO`` carrying the workbook bytes and a ``.name``, which is exactly
  the interface the page uses (``uploaded_file.read()``,
  ``uploaded_file.name``). Nothing else is patched, and no page logic is
  bypassed: the stub hands over bytes and the page does the rest.

  THE SEAM IS NOT A CONVENIENCE. ``AppTest`` grew a ``file_uploader`` widget
  accessor with ``set_value`` only in a recent release. This repository's CI
  matrix includes PYTHON 3.9, on which ``pip install ".[dev]"`` resolves
  ``streamlit==1.50.0`` -- whose ``streamlit.testing.v1.element_tree`` has no
  ``FileUploader`` class AT ALL. Measured, on 3.9.25, not assumed. A gate
  written against the widget accessor would therefore have needed a
  ``skipif`` on a quarter of the matrix, and a skip is how a gate stops
  asking. Stubbing the return value runs the IDENTICAL gate on every
  interpreter CI uses -- verified green on both 3.9/streamlit 1.50.0 and
  3.12/streamlit 1.62.0.

  IT DOES **NOT** OBSERVE:
    * Streamlit Cloud. This is a local in-process harness; it cannot run the
      deployed app, and the deployed app is what a CDE actually opens. A
      green here is a statement about the page's logic and emitted text, not
      about the hosted deployment;
    * the browser. No DOM, no CSS, no JavaScript, no layout, no fonts. What
      micromark does to these strings is modelled -- not executed -- by
      ``inline_math_spans`` in ``test_streamlit_markdown_survival.py``, whose
      own scope limit applies in full here;
    * any branch this drive does not enter. The assertions cover what one
      upload of one workbook renders. A string on a path not taken (a
      degraded eligibility fetch, an exception handler) is invisible here.
      ``TestTheOptimizerInfeasibilityStringKeepsItsUnits`` below covers one
      such branch as a PRODUCER-TO-RENDERER CONTRACT rather than a drive,
      and says why it cannot be a drive;
    * chart geometry, which is ``test_readiness_chart_geometry.py``.

  NETWORK. The drive calls ``Application.analyze``, which asks nmtc-mapper
  for eligibility data. When that download is unavailable the page takes its
  DEGRADED path and still renders every figure asserted below -- verified by
  running this drive with the socket layer hard-blocked and the on-disk cache
  hidden. Nothing here is conditional on the network, and nothing here skips
  when it is absent.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest

_REPO_ROOT = Path(__file__).resolve().parent.parent
_STREAMLIT_APP = _REPO_ROOT / "streamlit_app"
for _p in (str(_REPO_ROOT), str(_STREAMLIT_APP)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import openpyxl  # noqa: E402  (in the [dev] extra; a hard import, never a skip)
import streamlit as st  # noqa: E402
from streamlit.testing.v1 import AppTest  # noqa: E402

from tests.conftest import templates_dir  # noqa: E402
from tests.test_streamlit_markdown_survival import (  # noqa: E402
    inline_math_spans,
    rendered_text,
)

_PAGES = _STREAMLIT_APP / "pages"

#: What the CDE types into the two cells at issue. Deliberately NEITHER the
#: sample's own values NOR the placeholder: 48 is not 65, and CY2026 is not
#: blank, so a page that renders either of those is rendering something the
#: uploader did not supply.
_TYPED_ROUND = "CY2026"
_TYPED_ALLOCATION_MILLIONS = 48

#: The figure ``_UNSTATED_ALLOCATION_PLACEHOLDER`` puts in when an upload
#: states no allocation. It is correct for an upload that states none; it is
#: the DEFECT for an upload that states one, which is what this drives.
_PLACEHOLDER_RENDERINGS = ("65,000,000", "$65.0M")


def _cde_edited_workbook() -> bytes:
    """The shipped sample workbook with a CDE's OWN identity and two cells set.

    Read from the INSTALLED package, like every other template-reading gate
    here, so the sdist job -- which stages no ``nmtcapp/`` tree -- reads the
    workbook it actually ships rather than one from the working copy.

    The three identity cells are overwritten because
    ``core.sample_identity.assert_not_sample_identity`` REFUSES an upload that
    still carries the fictional Riverbend CDE's name, id and EIN. Changing one
    of the three does not clear it; that assumption is error #45 on this
    package's own ledger.
    """
    wb = openpyxl.load_workbook(Path(templates_dir()) / "pipeline_sample.xlsx")
    ws = wb["CDE Profile"]
    ws.cell(row=4, column=1).value = "Cascade Community Capital, LLC"
    ws.cell(row=4, column=2).value = "CDE-2020-0455"
    ws.cell(row=4, column=3).value = "45-9876543"
    ws.cell(row=4, column=10).value = _TYPED_ALLOCATION_MILLIONS
    ws.cell(row=4, column=11).value = _TYPED_ROUND
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class _UploadedFile(io.BytesIO):
    """Exactly the interface page 1 uses: ``.read()`` and ``.name``."""

    name = "cascade_pipeline.xlsx"


def _bodies(at: AppTest) -> list:
    """(kind, text) for every prose surface the page emitted."""
    out = []
    for kind in ("markdown", "caption", "info", "warning", "success", "error",
                 "header", "subheader", "title", "text"):
        for el in getattr(at, kind):
            value = getattr(el, "value", None)
            if isinstance(value, str):
                out.append((kind, value))
    return out


def _run_page1(workbook: bytes) -> AppTest:
    original = st.file_uploader
    st.file_uploader = lambda *a, **k: _UploadedFile(workbook)
    try:
        at = AppTest.from_file(str(_PAGES / "1_Pipeline_Analyzer.py"),
                               default_timeout=300)
        at.run()
        at.radio[0].set_value("Upload your own file").run()
        at.button[0].click().run()
    finally:
        st.file_uploader = original
    return at


def _run_downstream(page: str, carried: dict) -> AppTest:
    at = AppTest.from_file(str(_PAGES / page), default_timeout=300)
    for key, value in carried.items():
        at.session_state[key] = value
    at.run()
    return at


@pytest.fixture(scope="module")
def driven():
    """One upload, driven through pages 1, 2 and 3 as a CDE would walk them."""
    at1 = _run_page1(_cde_edited_workbook())
    carried = {
        k: at1.session_state[k]
        for k in ("app", "is_demo_data", "allocation_is_stated", "analysis")
        if k in at1.session_state
    }
    at2 = _run_downstream("2_Win_Alignment_Scorer.py", carried)
    at3 = _run_downstream("3_Pipeline_Optimizer.py", carried)
    return {"1": at1, "2": at2, "3": at3, "carried": carried}


# ---------------------------------------------------------------------------
# T2 -- the two facts survive the trip from the cell to the screen
# ---------------------------------------------------------------------------

class TestTheUploadersOwnFiguresReachTheScreen:
    def test_page_one_runs_the_upload_without_raising(self, driven):
        assert not driven["1"].exception, (
            "driving page 1 with an uploaded workbook raised: "
            + "; ".join(str(e.value) for e in driven["1"].exception)
        )

    def test_the_round_the_cde_typed_is_rendered(self, driven):
        text = "\n".join(v for _, v in _bodies(driven["1"]))
        assert _TYPED_ROUND in text, (
            f"page 1 never renders the round the CDE typed ({_TYPED_ROUND!r}). "
            "The CDE Profile sheet's 'Application Round' cell stated it and "
            "the page is showing something else -- through 1.5.6, "
            "'not specified - CDE to state'."
        )

    def test_the_allocation_the_cde_typed_is_rendered(self, driven):
        text = "\n".join(v for _, v in _bodies(driven["1"]))
        assert "$48.0M" in text, (
            "page 1 never renders the requested allocation the CDE typed "
            f"(${_TYPED_ALLOCATION_MILLIONS}.0M). The 'Requested Allocation "
            "($M)' cell stated it -- through 1.5.6 the page said 'not "
            "supplied [CDE TO COMPLETE]' instead."
        )

    @pytest.mark.parametrize("page", ["1", "2", "3"])
    def test_the_unstated_allocation_placeholder_is_rendered_nowhere(
        self, driven, page
    ):
        offenders = [
            f"{kind}: {value[:130]!r}"
            for kind, value in _bodies(driven[page])
            for placeholder in _PLACEHOLDER_RENDERINGS
            if placeholder in value
        ]
        assert not offenders, (
            f"page {page} renders the unstated-allocation placeholder to a "
            "CDE that STATED its allocation. Nothing on screen may show a "
            "money figure the uploader never typed:\n" + "\n".join(offenders)
        )


# ---------------------------------------------------------------------------
# T1b -- pages 2 and 3 inherit page 1's Application, so the defect did too
# ---------------------------------------------------------------------------

class TestTheDownstreamPagesInheritWhatPageOneBuilt:
    """``2_Win_Alignment_Scorer.py:65`` and ``3_Pipeline_Optimizer.py:99``
    both call ``get_or_create_app()`` with NO ARGUMENTS. Verified by driving
    them rather than by reading them: the object they score against is the
    SAME instance page 1 stored, so a wrong allocation on page 1 was a wrong
    allocation on all three -- and the fix on page 1 repairs all three.
    """

    @pytest.mark.parametrize("page", ["2", "3"])
    def test_the_page_reuses_the_identical_application_object(self, driven, page):
        assert driven[page].session_state["app"] is driven["1"].session_state["app"], (
            f"page {page} built its own Application instead of reusing the one "
            "page 1 stored. This gate's premise -- that the two downstream "
            "pages inherit page 1's round and allocation -- no longer holds, "
            "and their coverage above is asserting nothing."
        )

    @pytest.mark.parametrize("page", ["2", "3"])
    def test_the_inherited_application_carries_the_typed_figures(self, driven, page):
        app = driven[page].session_state["app"]
        assert app.application_round == _TYPED_ROUND
        assert app.requested_allocation == _TYPED_ALLOCATION_MILLIONS * 1_000_000


# ---------------------------------------------------------------------------
# T3 -- the money figures survive the markdown parser standing in front of them
# ---------------------------------------------------------------------------

class TestMoneyFiguresSurviveIntoTheTextAReaderSees:
    """The source strings were correct the whole time, which is why nothing
    saw this. ``completeness_check`` and ``consistency_check`` each emit TWO
    currency figures in ONE body; ``optimizer.constraints`` emits two more.
    Rendered raw, micromark's ``singleDollarTextMath`` pairs the two ``$``,
    eats both and re-typesets the run between them.
    """

    @pytest.mark.parametrize("page", ["1", "2", "3"])
    def test_no_rendered_body_opens_an_inline_math_span(self, driven, page):
        offenders = [
            f"{kind}: {value[:140]!r}"
            for kind, value in _bodies(driven[page])
            if inline_math_spans(value)
        ]
        assert not offenders, (
            f"page {page} emits a prose body whose '$' characters micromark "
            "would pair into an inline-math span, eating both delimiters and "
            "re-typesetting everything between them. Route it through "
            "utils.md():\n" + "\n".join(offenders)
        )

    def test_every_money_figure_on_page_one_keeps_its_dollar_sign(self, driven):
        losses = []
        for kind, value in _bodies(driven["1"]):
            seen = rendered_text(value)
            for figure in ("30,200,000", "48,000,000"):
                if figure in value and f"${figure}" not in seen:
                    losses.append(f"{kind}: {figure} loses its '$' in {value[:120]!r}")
        assert not losses, (
            "a money figure reaches the reader without its unit:\n"
            + "\n".join(losses)
        )


# ---------------------------------------------------------------------------
# T3, continued -- one two-dollar producer this drive cannot reach
# ---------------------------------------------------------------------------

class TestTheOptimizerInfeasibilityStringKeepsItsUnits:
    """``OptimizationConstraints.is_feasible`` emits "Total QEI $X < minimum
    $Y" and "... > maximum $Y" -- two currency figures in ONE body, the same
    shape as the validation warnings -- and
    ``3_Pipeline_Optimizer.py`` renders ``infeasibility_reason`` into an
    ``st.warning``.

    WHY THIS IS A CONTRACT TEST AND NOT A PAGE DRIVE, STATED PLAINLY. The
    branch is NOT reachable from page 3's own widgets, which was measured
    rather than assumed: the page constructs ``OptimizationConstraints`` with
    ``max_total_qei``, ``min_states``, ``min_projects`` and
    ``required_sectors`` only, and ``min_total_qei`` therefore keeps its
    default of ``0.0``, so the "<" branch cannot fire; driving the page across
    the slider and number-input ranges produced "5 projects < minimum 8" --
    a reason carrying no ``$`` at all -- or a feasible result.

    So the ``md()`` call on that line is DEFENSIVE, and this gate is honest
    about that: the defect there is LATENT, not live. It is gated anyway
    because the producer is one edit away from being reachable (any widget
    that sets ``min_total_qei`` makes it so), and because a reader of the diff
    is owed a test that says why the call is there.
    """

    def test_the_real_producers_two_figures_survive_md(self):
        from nmtcapp.core.pipeline import Pipeline
        from nmtcapp.optimizer.constraints import OptimizationConstraints
        from utils import md

        projects = list(Pipeline.sample(n=3))
        constraints = OptimizationConstraints(min_total_qei=900_000_000)
        ok, reason = constraints.is_feasible(projects)
        assert not ok and reason.count("$") == 2, (
            "this gate's premise is that is_feasible emits two currency "
            f"figures in one string; it emitted {reason!r}"
        )

        # The page's own body, verbatim from 3_Pipeline_Optimizer.py.
        body = md(
            f"\u26a0\ufe0f **Constraints not fully satisfied:** {reason}. "
            "The optimizer returned the best feasible result given the pipeline."
        )
        assert not inline_math_spans(body), (
            "the optimizer's infeasibility notice still opens an inline-math "
            f"span after md(): {body[:140]!r}"
        )
        seen = rendered_text(body)
        for figure in ("$20,500,000", "$900,000,000"):
            assert figure in seen, (
                f"{figure} loses its '$' on the way to the reader: {seen[:140]!r}"
            )
