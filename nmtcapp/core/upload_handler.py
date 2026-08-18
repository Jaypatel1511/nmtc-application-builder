"""File upload parsing for Pipeline Analyzer — handles xlsx (v1.1 template) and CSV."""
from __future__ import annotations

import io
import logging
import tempfile
from typing import Optional

import pandas as pd

from nmtcapp.core.pipeline import Pipeline
from nmtcapp.renderers._cell_format import NOT_SUPPLIED_INPUT

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# xlsx Pipeline sheet — display name → internal snake_case column map
# ---------------------------------------------------------------------------

_XLSX_PIPELINE_COL_MAP: dict[str, str] = {
    "Project ID":                   "project_id",
    "Project Name":                 "project_name",
    "QALICB Name":                  "qalicb_name",
    "Address":                      "address",
    "City":                         "city",
    "State":                        "state",
    "Sector":                       "sector",
    "Project Type":                 "project_type",
    "QEI ($M)":                     "qei_request",         # multiplied × 1_000_000 below
    "Total Cost ($M)":              "total_project_cost",   # multiplied × 1_000_000 below
    "Jobs Created":                 "expected_jobs_created",
    "Jobs Retained":                "expected_jobs_retained",
    "Affordable Units":             "expected_units_built",
    "Commercial Sq Ft":             "expected_sq_ft",
    "Distress Level":               "distress_level",
    "Native Area (Y/N)":            "native_area",
    "High Migration Rural (Y/N)":   "high_migration_rural",
    "US Territory (Y/N)":           "us_territory",
    "Persistent Poverty (Y/N)":     "persistent_poverty",
    "Below-Market Rate (Y/N)":      "below_market_rate",
    "Unrelated Entity (Y/N)":       "unrelated_entity",
    "Opportunity Zone (Y/N)":       "opportunity_zone",
    "Urban/Rural":                  "urban_rural",
    "Census Tract (11-digit)":      "census_tract",
    "Closing Target Date":          "closing_target_date",
    "Minority Owned (Y/N)":         "minority_owned",
    "Women Owned (Y/N)":            "women_owned",
    "Notes":                        "notes",
}

# After the rename, these columns need × 1_000_000 (template stores values in millions)
_XLSX_MILLIONS_COLS: frozenset[str] = frozenset({"qei_request", "total_project_cost"})

# ---------------------------------------------------------------------------
# CDE Profile sheet — display label → scoring attr key map
# ---------------------------------------------------------------------------

_CDE_FIELD_MAP: dict[str, str] = {
    "CDE Name":                             "cde_name",
    "CDE ID":                               "cde_id",
    "EIN":                                  "ein",
    "HQ State":                             "headquarters_state",
    "Certification Date":                   "certification_date",
    "Mission Statement":                    "mission",
    "Website":                              "website",
    "Org Type":                             "organization_type",
    "Target Markets (states, comma-sep)":   "target_markets",
    "Requested Allocation ($M)":            "requested_allocation_millions",
    "Application Round":                    "application_round",
    "Below-Market Rate Pct (0–1)*":    "products_below_market_pct",
    "Flexible Product Indicia Count":       "products_flexible_indicia_count",
    "Pipeline % Identified (0–1)":     "pipeline_pct_identified",
    "Own Capital at Risk (Y/N)":            "has_own_capital_at_risk",
    "Prior Award Count":                    "prior_award_count",
    "Years in Operation":                   "years_in_operation",
    "Track Record Pipeline Align (0–1)": "track_record_pipeline_alignment_pct",
    "Track Record Deployment Pct (0–1)": "track_record_deployment_pct",
    "Persistent Poverty Pct (0–1)*":   "pct_persistent_poverty",
    "US Territories Pct (0–1)*":       "pct_us_territories",
    "Quantified Outcomes (Y/N)":            "has_quantified_outcomes",
    "Third-Party Validation (Y/N)":         "has_third_party_validation",
    "LIC Board Representation (0–1)":  "lic_board_representation_pct",
    "Community Engagement Track Record (Y/N)": "has_community_engagement_track_record",
    "DBC Focus Years":                      "dbc_focus_years",
    "DBC Dollar Volume Pct (0–1)*":    "dbc_dollar_volume_pct",
    "Unrelated Entities Pct (0–1)*":   "unrelated_entities_pct",
    "Favorable Fee Structure (Y/N/Unknown)": "has_favorable_fee_structure",
    "Prior Reporting Issues (Y/N)":         "has_prior_reporting_issues",
}

_BOOL_FIELDS: frozenset[str] = frozenset({
    "has_own_capital_at_risk", "has_quantified_outcomes", "has_third_party_validation",
    "has_community_engagement_track_record", "has_prior_reporting_issues",
    "has_favorable_fee_structure",
})


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _read_csv_with_encoding_fallback(buf: io.BytesIO) -> pd.DataFrame:
    """Try utf-8 → utf-8-sig → cp1252 → latin-1 when reading a CSV."""
    for encoding in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            buf.seek(0)
            df = pd.read_csv(buf, encoding=encoding)
            if encoding not in ("utf-8", "utf-8-sig"):
                logger.warning(
                    "CSV decoded with %s (not UTF-8). "
                    "To avoid this warning, re-save the file as UTF-8.",
                    encoding,
                )
            return df
        except UnicodeDecodeError:
            continue
    raise ValueError(
        "Could not decode the CSV file. Try re-saving it with UTF-8 encoding "
        "(in Excel: File → Save As → CSV UTF-8)."
    )


def _read_pipeline_sheet_from_wb(wb, sheet_name: str) -> pd.DataFrame:
    """Read the Pipeline sheet from an openpyxl workbook.

    The v1.1 template has a 3-row preamble (title / section banners / column headers)
    before data rows.  Plain xlsx files (column headers in row 1) are also supported.
    """
    ws = wb[sheet_name]
    cell_a1_val = ws.cell(1, 1).value
    if cell_a1_val and "NMTC Application Builder" in str(cell_a1_val):
        header_row = 3
        data_start_row = 4
    else:
        header_row = 1
        data_start_row = 2

    max_col = ws.max_column
    headers = [ws.cell(header_row, c).value for c in range(1, max_col + 1)]

    rows = []
    for r in range(data_start_row, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        if any(v is not None for v in row_vals):
            rows.append(row_vals)

    return pd.DataFrame(rows, columns=headers)


def _parse_cde_profile_from_wb(wb) -> Optional[dict]:
    """Extract CDE-level scoring attrs from the 'CDE Profile' sheet.

    Template layout: row 1 = title, row 2 = section banners, row 3 = field headers, row 4 = data.
    """
    try:
        ws = wb["CDE Profile"]
        headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
        data_row = [ws.cell(4, c).value for c in range(1, ws.max_column + 1)]
    except Exception:
        return None

    result: dict = {}
    for col_idx, header in enumerate(headers):
        if header is None or (isinstance(header, float) and pd.isna(header)):
            continue
        header_str = str(header).strip()
        key = _CDE_FIELD_MAP.get(header_str)
        if key is None:
            continue
        val = data_row[col_idx] if col_idx < len(data_row) else None
        if val is None or (isinstance(val, float) and pd.isna(val)):
            continue
        if key in _BOOL_FIELDS:
            val_str = str(val).strip().upper()
            result[key] = val_str in ("Y", "YES", "TRUE", "1")
        elif key in ("products_flexible_indicia_count", "prior_award_count",
                     "years_in_operation", "dbc_focus_years"):
            try:
                result[key] = int(val)
            except (ValueError, TypeError):
                pass
        elif key in ("products_below_market_pct", "pipeline_pct_identified",
                     "track_record_pipeline_alignment_pct", "track_record_deployment_pct",
                     "pct_persistent_poverty", "pct_us_territories",
                     "lic_board_representation_pct", "dbc_dollar_volume_pct",
                     "unrelated_entities_pct"):
            try:
                result[key] = float(val)
            except (ValueError, TypeError):
                pass
        else:
            result[key] = str(val).strip()

    return result if result else None


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def load_uploaded_pipeline(
    file_bytes: bytes, filename: str
) -> tuple[Pipeline, Optional[dict]]:
    """Parse an uploaded CSV or Excel file into a Pipeline and optional CDE metadata.

    Returns:
        (Pipeline, cde_extra_or_None) — cde_extra is populated when the file
        contains a 'CDE Profile' sheet with v1.1 CDE-level scoring fields.
    """
    ext = filename.rsplit(".", 1)[-1].lower()
    buf = io.BytesIO(file_bytes)
    cde_extra: Optional[dict] = None

    if ext == "csv":
        df = _read_csv_with_encoding_fallback(buf)

    elif ext in ("xlsx", "xls"):
        import openpyxl as _openpyxl
        wb = _openpyxl.load_workbook(buf, data_only=True)

        if "CDE Profile" in wb.sheetnames:
            cde_extra = _parse_cde_profile_from_wb(wb)

        sheet_name = "Pipeline" if "Pipeline" in wb.sheetnames else wb.sheetnames[0]
        df = _read_pipeline_sheet_from_wb(wb, sheet_name)

        # Drop None column names (merged-cell artefacts from the template header rows)
        df = df.loc[:, [c for c in df.columns if c is not None]]

        # Map display names → internal snake_case names
        df = df.rename(columns={k: v for k, v in _XLSX_PIPELINE_COL_MAP.items()
                                 if k in df.columns})

        # Convert $M columns to raw dollars
        for col in _XLSX_MILLIONS_COLS:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0) * 1_000_000

    else:
        raise ValueError(f"Unsupported file type: .{ext}")

    # ── Synthesise missing required columns ──────────────────────────────────
    if "qalicb_name" not in df.columns:
        df["qalicb_name"] = df.get("project_name", "")
    if "address" not in df.columns:
        df["address"] = ""
    # THE FIGURE THE CDE NEVER SUPPLIED (1.3.0 S3).
    #
    # These two lines used to be the whole of it: no flag, no warning, no trace
    # downstream. The synthesised value then rendered as "Total QLICI ($)" in
    # Appendix A — the CDE's own answer to the CDFI Fund's Table A5 row (h) —
    # on markdown, Word, PDF and Excel, and satisfied validation's QLICI <= QEI
    # rule trivially by being exactly equal to the QEI it was copied from.
    #
    # The default itself STAYS. qlici_amount is non-optional on
    # PipelineProject, is in REQUIRED_PROJECT_FIELDS, and several downstream
    # consumers dereference it; removing the value would turn a missing column
    # into an exception on upload, which is a bigger behaviour change than this
    # release is scoped for and is not what a CDE who omitted one column needs.
    # What changes is that the document no longer PRESENTS it. The flag rides
    # the temp CSV below into Pipeline.from_csv, and every surface that prints
    # the figure reads the flag first.
    if "qlici_amount" not in df.columns and "qei_request" in df.columns:
        df["qlici_amount"] = df["qei_request"]
        df["qlici_amount_supplied"] = "N"
        # WARN HERE TOO. The silent default is what hid this for four passes,
        # and a reader of the rendered document sees the consequence without
        # seeing the cause. The Streamlit uploader surfaces the same fact from
        # the carried flag (streamlit_app/pages/1_Pipeline_Analyzer.py).
        logger.warning(
            "The uploaded file has no 'qlici_amount' column. Each project's "
            "QEI request has been used in its place so the pipeline can load, "
            "but that figure is NOT the CDE's QLICI amount and is not "
            "presented as one: it renders as %r wherever the document would "
            "print Total QLICI, and the QLICI <= QEI consistency check is "
            "reported as not checkable rather than passed. Supply a "
            "qlici_amount column before filing.",
            NOT_SUPPLIED_INPUT,
        )

    # ── Pipeline-derived CDE pcts from per-project Y/N flags ─────────────────
    if "qei_request" in df.columns:
        cde_extra = cde_extra or {}
        total_qei = pd.to_numeric(df["qei_request"], errors="coerce").fillna(0).sum()
        if total_qei > 0:
            for flag_col, attr_key in (
                ("below_market_rate", "products_below_market_pct"),
                ("unrelated_entity",  "unrelated_entities_pct"),
                ("us_territory",      "pct_us_territories"),
                ("persistent_poverty", "pct_persistent_poverty"),
            ):
                if flag_col in df.columns and attr_key not in cde_extra:
                    flag_qei = (
                        df[df[flag_col].astype(str).str.strip().str.upper() == "Y"][
                            "qei_request"
                        ]
                        .pipe(pd.to_numeric, errors="coerce")
                        .fillna(0)
                        .sum()
                    )
                    cde_extra[attr_key] = round(flag_qei / total_qei, 4)

    with tempfile.NamedTemporaryFile(
        suffix=".csv", delete=False, mode="w", newline=""
    ) as tmp:
        df.to_csv(tmp, index=False)
        tmp_path = tmp.name

    return Pipeline.from_csv(tmp_path), cde_extra
