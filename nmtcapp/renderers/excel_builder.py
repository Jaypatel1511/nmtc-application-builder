"""Professional Excel workbook builder for NMTC applications."""
from __future__ import annotations

import logging
import os
from datetime import date
from typing import TYPE_CHECKING

from nmtcapp.data.schema import READINESS_SCORING_WEIGHTS
from nmtcapp.renderers._cell_format import is_identifier_column
from nmtcapp.renderers._disclosure import (
    is_partial_unverified, qualified_pct, unverified_banner,
)
from nmtcapp.renderers._methodology import (
    readiness_inline_qualifier,
    readiness_weights_sheet_note,
)
from nmtcapp.renderers._round_provenance import (
    ROUND_PROVENANCE_SHEET_NAME, round_provenance_paragraphs,
)
from nmtcapp.renderers._question_25 import (
    Q25_BASIS_LABEL, Q25_BASIS_SHEET_NAME, Q25_QEI_BASIS_SUFFIX_SHEET,
    q25_basis_note_paragraphs,
)
from nmtcapp.renderers._sheet_geometry import (
    MAX_ROW_HEIGHT, required_row_height, span_points,
)
from nmtcapp.renderers.styles import COLORS, TYPOGRAPHY, TABLE_STYLES, xl_color
from nmtcapp.tables.distress_table import build_distress_table
from nmtcapp.tables.geographic_table import build_geographic_table
from nmtcapp.tables.impact_table import build_impact_table
from nmtcapp.tables.investor_table import build_investor_table
from nmtcapp.tables.investor_table import INVESTOR_TABLE_TITLE
from nmtcapp.tables.pipeline_table import build_pipeline_table
from nmtcapp.tables.track_record_table import build_track_record_table

if TYPE_CHECKING:
    from nmtcapp.core.application import Application, ApplicationAnalysis

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


# ---------------------------------------------------------------------------
# Style constants (openpyxl PatternFill / Font objects built on first use)
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> "PatternFill":
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font(bold=False, color="1A1A1A", size=10, name="Calibri", italic=False) -> "Font":
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)


def _center() -> "Alignment":
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> "Alignment":
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _right() -> "Alignment":
    return Alignment(horizontal="right", vertical="center")


def _thin_border() -> "Border":
    thin = Side(style="thin", color=xl_color("border"))
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# Number formats
FMT_CURRENCY = '"$"#,##0'
FMT_CURRENCY_DEC = '"$"#,##0.00'
FMT_PCT = "0.0%"
FMT_PCT0 = "0%"
FMT_NUMBER = "#,##0"
FMT_DECIMAL2 = "0.00"
FMT_TEXT = "@"
# Digits, no separator, no decimal point — for a value that is a LABEL rather
# than a quantity (1.2.1 B-3). "0" and not FMT_TEXT: the cell keeps its numeric
# type so a sort on the column still sorts by year, and Excel does not stamp it
# with the green "number stored as text" warning triangle.
FMT_IDENTIFIER = "0"


# --- Summary Dashboard geometry (1.3.0 B1) ---------------------------------
#
# The column widths were set at the BOTTOM of _build_summary_dashboard, forty
# lines after the last row that had to be sized against them. Every row height
# on the sheet therefore had to be a guess about a width the code had not set
# yet, and the guesses were hand-typed: 18 for every metrics row, 28 for the
# weights disclosure, and a _CHARS_PER_LINE constant for the basis note.
#
# They are declared here instead, once, ahead of everything that reads them,
# so a height is DERIVED from the width it has to fit rather than guessed
# before it exists. tests/test_excel_geometry.py re-derives both from the
# shipped file and fails when a label cannot display in its own row.
DASHBOARD_COL_WIDTHS = {
    "A": 22.0, "B": 18.0, "C": 18.0, "D": 12.0, "E": 12.0, "F": 12.0,
}

#: The metrics block merges A:B for its labels; the note rows merge A:F.
DASHBOARD_LABEL_SPAN_PTS = span_points(
    DASHBOARD_COL_WIDTHS["A"], DASHBOARD_COL_WIDTHS["B"],
)
DASHBOARD_FULL_SPAN_PTS = span_points(*DASHBOARD_COL_WIDTHS.values())


class ExcelApplicationBuilder:
    """Builds a professional Excel workbook for the NMTC application package.

    Produces 7 sheets: Pipeline Detail, Distress Documentation, Geographic
    Targeting, Impact Projections, Investor Commitments, Track Record, and a
    Summary Dashboard.

    Example::

        builder = ExcelApplicationBuilder(application, analysis)
        builder.save("./drafts/application.xlsx")
    """

    def __init__(self, application: "Application", analysis: "ApplicationAnalysis") -> None:
        if not _OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel output. Install with: pip install openpyxl")
        self.application = application
        self.analysis = analysis

    def build(self) -> "Workbook":
        """Build and return an openpyxl Workbook.

        Example::

            wb = builder.build()
        """
        wb = Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        self._build_summary_dashboard(wb)
        # Second tab on purpose: the Summary Dashboard's distress label names
        # this sheet, and a pointer is only as good as how far the reader has
        # to look for what it names.
        self._build_q25_basis_sheet(wb)
        # THIRD TAB, AND THE POSITION IS THE FIX (1.5.0 B1). Until this
        # release the workbook was the ONLY one of the four formats carrying
        # no round provenance at all: markdown, Word and PDF each stated that
        # CY 2024-2025 is closed and awarded and that CY 2026 is unpublished,
        # while the workbook cited that round in the present tense on the Q25
        # Basis Note sheet and said nothing about which round it was. Excel is
        # the format most likely to be circulated internally and pasted from,
        # so it was the worst one to leave silent.
        self._build_round_provenance_sheet(wb)
        self._build_pipeline_sheet(wb)
        self._build_distress_sheet(wb)
        self._build_geographic_sheet(wb)
        self._build_impact_sheet(wb)
        self._build_investor_sheet(wb)
        self._build_track_record_sheet(wb)

        return wb

    def save(self, path: str) -> None:
        """Save the workbook to an .xlsx file.

        Example::

            builder.save("./output/application.xlsx")
        """
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        wb = self.build()
        wb.save(path)
        size_kb = os.path.getsize(path) // 1024
        logger.info("Excel workbook saved: %s (%d KB)", path, size_kb)

    # ------------------------------------------------------------------
    # Sheet 1: Summary Dashboard
    # ------------------------------------------------------------------

    def _build_summary_dashboard(self, wb: "Workbook") -> None:
        ws = wb.create_sheet("Summary Dashboard")
        ws.sheet_view.showGridLines = False
        # Widths first: every row height below is derived from the span it has
        # to fit, and the span does not exist until the widths are set.
        for col, width in DASHBOARD_COL_WIDTHS.items():
            ws.column_dimensions[col].width = width
        app = self.application
        analysis = self.analysis
        pr = analysis.pipeline_result
        score = analysis.readiness_score

        # Title block
        ws.merge_cells("A1:F1")
        ws["A1"] = "NEW MARKETS TAX CREDIT ALLOCATION APPLICATION"
        ws["A1"].font = _font(bold=True, color="FFFFFF", size=16)
        ws["A1"].fill = _fill(xl_color("primary"))
        ws["A1"].alignment = _center()
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:F2")
        ws["A2"] = app.cde.name
        ws["A2"].font = _font(bold=True, color="FFFFFF", size=13)
        ws["A2"].fill = _fill(xl_color("secondary"))
        ws["A2"].alignment = _center()
        ws.row_dimensions[2].height = 22

        degraded = getattr(pr, "eligibility_data_status", "ok") != "ok"
        partial_unverified = is_partial_unverified(pr)
        ws.merge_cells("A3:F3")
        # THE FIFTH COVER CLAIM, AND EXCEL WAS SUPPOSED TO BE THE GOOD SURFACE
        # (1.5.1 audit, F3). T3 added readiness_inline_qualifier() to the Word,
        # PDF and Markdown cover tables and left this one alone, on the
        # reasoning that "Excel was the only surface that put the disclosure
        # where the claim was" -- measured at 48 characters.
        #
        # That measurement was taken with the OLD proximity matcher, which
        # accepted any generic house disclaimer as any claim's disclosure. Under
        # a matcher that requires the READINESS disclosure specifically, this
        # banner's nearest one is 633 characters away AND ON A DIFFERENT SHEET:
        # readiness_weights_sheet_note() lives on the breakdown tab, and a
        # reader looking at the dashboard never sees it. The 48-character figure
        # was the distance to a disclosure about something else.
        #
        # So the surface held up as the model for the other three was the one
        # with an undisclosed grade on its front page.
        ws["A3"] = (
            f"{app.application_round}  |  Prepared: {date.today().strftime('%B %d, %Y')}  |  "
            f"Readiness Grade: {score.grade} ({score.overall_score:.1f}/100"
            + (" PARTIAL)" if getattr(score, "partial", False) else ")")
            + f" — {readiness_inline_qualifier()}"
        )
        ws["A3"].font = _font(color="FFFFFF", size=9, italic=True)
        ws["A3"].fill = _fill(xl_color("accent"))
        ws["A3"].alignment = _center()
        # THE HEIGHT IS DERIVED, like the banner below and for the same reason
        # (1.3.0 B1's class). This was a hardcoded ``16``, which fitted the
        # banner only while the banner was short. Adding the readiness
        # qualifier above pushed it to two lines, and a hardcoded 16 would have
        # clipped the disclosure -- shipping an undisclosed grade in a way that
        # looks exactly like a disclosed one in the source.
        #
        # Clamped on write like every other row here; test_excel_geometry
        # asserts against the UNCLAMPED value, so reaching the ceiling fails
        # loudly instead of truncating silently.
        ws.row_dimensions[3].height = min(
            MAX_ROW_HEIGHT,
            required_row_height(ws["A3"].value, DASHBOARD_FULL_SPAN_PTS, 9),
        )

        # THE BANNER ROW READS ITS OWN TEXT (1.3.0 FIX-2 B2). Both arms below
        # hardcoded ``height = 28`` — in this function, two rows above the cell
        # 1.3.0 B1 fixed by deriving its height, and left behind by that fix.
        # Measured on the shipped builder at the shipped geometry:
        #
        #     partial-unverified   ships 28.0 pt, needs 75.0 pt (519 chars)
        #     degraded             ships 28.0 pt, needs 30.0 pt (168 chars)
        #
        # The degraded arm clips at its SHORTEST possible message, so no
        # wording makes 28 correct. What the partial banner never displayed:
        # "... so each is a LOWER BOUND ... Do not submit until all project
        # locations are verified." The row that exists to stop a CDE filing was
        # the row that could not finish its own sentence.
        #
        # A banner is the one row a reader must not have to widen a column to
        # read, and it is the row whose length is least predictable — it names
        # every unverified project, so it grows with the pipeline. Nothing here
        # may be a literal again.
        banner_text = None
        if degraded:
            banner_text = (
                "ELIGIBILITY DATA UNAVAILABLE — "
                f"{getattr(pr, 'eligibility_data_error', None) or 'reason unknown'}. "
                "Eligibility/distress figures are unverified; readiness score is "
                "partial (computed without eligibility verification)."
            )
        elif partial_unverified:
            banner_text = "UNVERIFIED PROJECT LOCATIONS — " + unverified_banner(pr)
        if banner_text is not None:
            ws.merge_cells("A4:F4")
            ws["A4"] = banner_text
            ws["A4"].font = _font(bold=True, color="FFFFFF", size=10)
            ws["A4"].fill = _fill("B00000")
            ws["A4"].alignment = _center()
            # Clamped, like the Q25 note block below: required_row_height does
            # not clamp on purpose, and a banner past Excel's 409 pt ceiling
            # would need splitting across rows by a person. The gate asserts
            # against the UNCLAMPED value, so it fails loudly if that ever
            # happens rather than shipping a silently truncated warning.
            ws.row_dimensions[4].height = min(
                MAX_ROW_HEIGHT,
                required_row_height(banner_text, DASHBOARD_FULL_SPAN_PTS, 10),
            )

        # --- Key Metrics block (rows 5–14) ---
        ws["A5"] = "KEY APPLICATION METRICS"
        ws["A5"].font = _font(bold=True, color=xl_color("primary"), size=11)
        ws.merge_cells("A5:F5")

        distress = pr.distress_breakdown
        impact = pr.aggregate_impact
        metrics = [
            ("Total NMTC Allocation Requested", app.requested_allocation, FMT_CURRENCY, "primary"),
            ("Total QEI in Pipeline", pr.total_qei_request, FMT_CURRENCY, "secondary"),
            ("Total Project Cost", pr.total_project_cost, FMT_CURRENCY, "secondary"),
            ("Number of Projects", pr.total_projects, FMT_NUMBER, "secondary"),
            ("States Represented", pr.geographic_diversity.get("states_count", 0), FMT_NUMBER, "secondary"),
            ("NMTC Eligibility Rate",
             ("Unverified" if degraded
              else qualified_pct(pr.eligibility_pct, pr) if partial_unverified
              else pr.eligibility_pct),
             FMT_TEXT if (degraded or partial_unverified) else FMT_PCT, None),
            # THE CELL WITH NO DENOMINATOR (1.3.0 S4). This label read
            # "Deep/Severe Distress Concentration" over a raw float under a
            # percent format, with no denominator in the label and no basis
            # note anywhere in the workbook — the one artifact of the four that
            # carried neither half of the 1.2.1 remedy. A CDE copying this cell
            # into Question 25 files a QEI figure against a QLICI commitment.
            #
            # It was invisible to the rendered baseline because the baseline
            # stores it as `|float|fmt=0.0%|0.8531...` — a number and a format
            # code, with the label on the row above and no text of its own to
            # scan. Naming the basis in the LABEL is what moves that entry's
            # neighbour; the basis note below is what gives the reader the rest.
            ("Deep/Severe Distress Concentration " + Q25_QEI_BASIS_SUFFIX_SHEET,
             ("Unverified" if degraded
              else qualified_pct(distress.get("pct_deep_or_severe", 0), pr)
              if partial_unverified
              else distress.get("pct_deep_or_severe", 0)),
             FMT_TEXT if (degraded or partial_unverified) else FMT_PCT, None),
            ("Total Jobs to Be Created", impact.get("total_jobs_created", 0), FMT_NUMBER, None),
            ("Jobs per $1MM QEI", impact.get("jobs_per_million_qei", 0), FMT_DECIMAL2, None),
        ]

        for i, (label, val, fmt, color_key) in enumerate(metrics):
            row = 6 + i
            # THE UNIFORM 18 IS WHAT B1 WAS. Every row of this loop got a
            # one-line height, and 1.3.0 S4 gave row 12 a label that needs
            # two — measured in Excel 16.112 at 30.0 pt against this exact
            # span. The label displayed as far as "(a share of QEI, not" and
            # stopped, so the visible half of the sentence ended mid-negation
            # and the half that never displayed was the pointer to the note.
            # A height that does not read its own label is a height that
            # cannot survive the label being edited, which is the failure
            # this loop already had once.
            ws.row_dimensions[row].height = required_row_height(
                label, DASHBOARD_LABEL_SPAN_PTS, 10,
            )
            label_cell = ws.cell(row=row, column=1, value=label)
            val_cell = ws.cell(row=row, column=3, value=val)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

            bg = xl_color("row_even") if i % 2 == 0 else xl_color("row_odd")
            label_cell.fill = _fill(bg)
            val_cell.fill = _fill(bg)
            label_cell.font = _font(size=10)
            label_cell.alignment = _left()
            label_cell.border = _thin_border()
            val_cell.number_format = fmt
            val_cell.font = _font(bold=True, size=10)
            val_cell.alignment = _right()
            val_cell.border = _thin_border()

        # --- Readiness Score block (rows 16–22) ---
        ws["A16"] = "READINESS SCORE BREAKDOWN"
        ws["A16"].font = _font(bold=True, color=xl_color("primary"), size=11)
        ws.merge_cells("A16:F16")

        ws.cell(row=17, column=1, value="Component").font = _font(bold=True, color="FFFFFF", size=10)
        ws.cell(row=17, column=3, value="Score").font = _font(bold=True, color="FFFFFF", size=10)
        ws.cell(row=17, column=4, value="Weight").font = _font(bold=True, color="FFFFFF", size=10)
        for col in (1, 2, 3, 4):
            ws.cell(row=17, column=col).fill = _fill(xl_color("primary"))
            ws.cell(row=17, column=col).alignment = _center()
            ws.cell(row=17, column=col).border = _thin_border()
        ws.merge_cells("A17:B17")
        ws.row_dimensions[17].height = 18

        # READ THE WEIGHTS, DO NOT RETYPE THEM. This was a hardcoded dict whose
        # keys were "eligibility" and "validation" while ReadinessScore's
        # component keys are "eligibility_quality" and "validation_pass_rate",
        # so weight_map.get(comp, 0) returned 0 for both. The workbook the Word
        # and PDF documents cross-reference as the authoritative attachment
        # printed Eligibility Quality 0.0% and Validation Pass Rate 0.0%, a
        # Weight column summing to 65%, and two components declared weightless
        # — against a methodology appendix in the same package stating 25% and
        # 10%. Reading the constant makes the two agree by construction and
        # removes the key-drift failure mode entirely.
        weight_map = READINESS_SCORING_WEIGHTS
        for i, (comp, val) in enumerate(score.component_scores.items()):
            row = 18 + i
            ws.row_dimensions[row].height = 17
            bg = xl_color("row_even") if i % 2 == 0 else xl_color("row_odd")

            label_cell = ws.cell(row=row, column=1, value=comp.replace("_", " ").title())
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            score_cell = ws.cell(row=row, column=3, value=val)
            weight_cell = ws.cell(row=row, column=4, value=weight_map.get(comp, 0))

            for cell in (label_cell, score_cell, weight_cell):
                cell.fill = _fill(bg)
                cell.border = _thin_border()
            label_cell.font = _font(size=10)
            label_cell.alignment = _left()
            score_cell.number_format = "0.0"
            score_cell.font = _font(bold=True, size=10)
            score_cell.alignment = _right()
            weight_cell.number_format = FMT_PCT
            weight_cell.font = _font(size=10)
            weight_cell.alignment = _right()

            # Conditional color: ≥75 green, ≥50 yellow, <50 red
            if val >= 75:
                score_cell.font = _font(bold=True, color=xl_color("green"), size=10)
            elif val >= 50:
                score_cell.font = _font(bold=True, color=xl_color("yellow"), size=10)
            else:
                score_cell.font = _font(bold=True, color=xl_color("red"), size=10)

        # Overall score totals row
        tot_row = 18 + len(score.component_scores)
        ws.row_dimensions[tot_row].height = 20
        overall_label = ws.cell(row=tot_row, column=1, value="OVERALL SCORE")
        ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=2)
        overall_val = ws.cell(row=tot_row, column=3, value=score.overall_score)
        overall_grade = ws.cell(row=tot_row, column=4, value=f"Grade {score.grade}")
        for cell in (overall_label, overall_val, overall_grade):
            cell.fill = _fill(xl_color("row_total"))
            cell.font = _font(bold=True, size=11)
            cell.border = _thin_border()
        overall_label.alignment = _left()
        overall_val.number_format = "0.0"
        overall_val.alignment = _right()
        overall_grade.alignment = _center()

        # The workbook showed a Weight column with no statement of whose
        # weighting it was — the only one of the four artifacts to print the
        # readiness breakdown without the disclosure the other three carry, and
        # the one a reviewer opens to check the numbers.
        disclosure_row = tot_row + 1
        ws.merge_cells(start_row=disclosure_row, start_column=1,
                       end_row=disclosure_row, end_column=6)
        disclosure_text = readiness_weights_sheet_note()
        ws.row_dimensions[disclosure_row].height = required_row_height(
            disclosure_text, DASHBOARD_FULL_SPAN_PTS, 8,
        )
        note_cell = ws.cell(row=disclosure_row, column=1, value=disclosure_text)
        note_cell.font = _font(italic=True, color=xl_color("text_muted"), size=8)
        note_cell.alignment = _left()

        # --- Bar chart: component scores ---
        try:
            score_start_row = 18
            score_end_row = 18 + len(score.component_scores) - 1
            chart = BarChart()
            chart.type = "bar"
            chart.title = "Readiness Score Components"
            chart.y_axis.title = "Score (0–100)"
            chart.style = 10
            chart.width = 14
            chart.height = 8

            data_ref = Reference(ws, min_col=3, min_row=score_start_row,
                                 max_row=score_end_row)
            cats_ref = Reference(ws, min_col=1, min_row=score_start_row,
                                 max_row=score_end_row)
            chart.add_data(data_ref)
            chart.set_categories(cats_ref)
            chart.series[0].title = SeriesLabel(v="Score")

            ws.add_chart(chart, "F5")
        except Exception:
            pass  # chart is optional — don't fail the build

        # THE BASIS NOTE IS NO LONGER ON THIS SHEET (1.3.0 B1).
        #
        # 1.3.0 S4 put it here, at rows 27-28, and pointed at it from the row
        # 12 label with the flowing-document suffix "see the basis note
        # below". Measured in Excel 16.112, that pointer was clipped — row 12
        # needed 30.0 pt and shipped 18.0 — and the note it pointed at was
        # fifteen rows past the readiness block, under a footer, at a row
        # number nothing on screen names.
        #
        # The height half is fixed above. The REACHABILITY half is not fixable
        # in place, and this is the finding that decided it: the workbook
        # stores no window size, no zoom and no frozen pane, so whether row 28
        # is "below and visible" or "below and off-screen" is a property of the
        # reader's monitor and not of this file. The audit's window showed
        # $A$1:$F$23 and the note was off-screen; the window this fix was
        # measured in showed $A$1:$X$28 and it was on-screen. Sizing a row
        # cannot repair a pointer whose truth varies by reader.
        #
        # So the note moves to a sheet the row 12 label names by name. A tab is
        # visible on open at every window size, and the name is a constant
        # rather than an arithmetic result — where "row 27" would have been a
        # sixth hand-typed count derived from len(component_scores).
        #
        # It is still the same string from the same function as Section B, so
        # a future correction still cannot land on three surfaces and miss the
        # fourth. Only its address changed.

        # Footer
        footer_row = disclosure_row + 2
        ws.merge_cells(f"A{footer_row}:F{footer_row}")
        ws.cell(row=footer_row, column=1,
                value=f"CONFIDENTIAL — {app.cde.name} — NMTC {app.application_round} — "
                      f"Generated {date.today().isoformat()}")
        ws.cell(row=footer_row, column=1).font = _font(italic=True, color=xl_color("text_muted"), size=8)
        ws.cell(row=footer_row, column=1).alignment = _center()

    # ------------------------------------------------------------------
    # Sheet 2: Q25 Basis Note
    # ------------------------------------------------------------------

    def _build_q25_basis_sheet(self, wb: "Workbook") -> None:
        """The basis note, on the sheet the dashboard's distress label names.

        WHY IT IS A SHEET (1.3.0 B1). It was two rows at the bottom of the
        Summary Dashboard, pointed at by "see the basis note below" — a
        direction whose truth depends on the reader's window size, because the
        workbook stores no window geometry at all. The label now names this
        sheet instead, and a tab is visible on open regardless of monitor.

        Every height here is derived from the text and the span it has to fit,
        not typed. The version this replaces set the note's row to
        ``11 * (len(text) // 130 + 2)`` — 275 pt against a measured 217, so
        58 pt of blank cell below the last line, from a guess that happened to
        err in the safe direction. The row above it, sized by the same kind of
        guess, erred in the other one.
        """
        ws = wb.create_sheet(Q25_BASIS_SHEET_NAME)
        ws.sheet_view.showGridLines = False
        for col, width in DASHBOARD_COL_WIDTHS.items():
            ws.column_dimensions[col].width = width

        ws.merge_cells("A1:F1")
        heading = ws.cell(row=1, column=1, value=Q25_BASIS_LABEL)
        heading.font = _font(bold=True, color=xl_color("primary"), size=11)
        heading.alignment = _left()
        ws.row_dimensions[1].height = required_row_height(
            Q25_BASIS_LABEL, DASHBOARD_FULL_SPAN_PTS, 11,
        )

        # Names the cell it qualifies, so the pointer works in both directions:
        # the dashboard says which sheet, this sheet says which cell.
        pointer = (
            "This note is the basis for 'Deep/Severe Distress Concentration' "
            "on the Summary Dashboard, and for the Deep Distress and Severely "
            "Distressed rows of Section B."
        )
        ws.merge_cells("A2:F2")
        pointer_cell = ws.cell(row=2, column=1, value=pointer)
        pointer_cell.font = _font(color=xl_color("text_muted"), size=9)
        pointer_cell.alignment = _left()
        ws.row_dimensions[2].height = required_row_height(
            pointer, DASHBOARD_FULL_SPAN_PTS, 9,
        )

        # ONE PARAGRAPH PER ROW, not one cell (1.3.0 B1). As a single merged
        # cell this note needed 405 pt against Excel's 409-pt ceiling — three
        # points of headroom on text that has grown in three of the last four
        # rounds. The next sentence added to it would have clipped silently
        # with every height check still passing.
        #
        # required_row_height does not clamp, so if any single paragraph ever
        # outgrows one row, tests/test_excel_geometry.py fails on it rather
        # than the workbook truncating it.
        for offset, para in enumerate(q25_basis_note_paragraphs()):
            row = 4 + offset
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value=para)
            cell.font = _font(size=9)
            cell.alignment = Alignment(horizontal="left", vertical="top",
                                       wrap_text=True)
            ws.row_dimensions[row].height = min(
                MAX_ROW_HEIGHT,
                required_row_height(para, DASHBOARD_FULL_SPAN_PTS, 9),
            )

    def _build_round_provenance_sheet(self, wb: "Workbook") -> None:
        """WHICH ROUND this workbook is based on -- the B1 fix.

        The text is not composed here. It is
        ``_round_provenance.round_provenance_paragraphs()``, whose join is the
        exact string Word, PDF and markdown render, so the workbook cannot
        drift from the other three formats and a sentence cannot be corrected
        in one place and left stale here. That drift is the 1.2.0 defect this
        module's source-of-truth rule exists to prevent.

        One paragraph per row for the same reason the Q25 basis note is: a
        merged cell has a 409-pt ceiling and this note is longer than that, so
        a single-cell version would clip silently with every height check
        still passing.
        """
        ws = wb.create_sheet(ROUND_PROVENANCE_SHEET_NAME)
        ws.sheet_view.showGridLines = False
        for col, width in DASHBOARD_COL_WIDTHS.items():
            ws.column_dimensions[col].width = width

        heading_text = "WHICH ROUND THIS WORKBOOK IS BASED ON"
        ws.merge_cells("A1:F1")
        heading = ws.cell(row=1, column=1, value=heading_text)
        heading.font = _font(bold=True, color=xl_color("primary"), size=11)
        heading.alignment = _left()
        ws.row_dimensions[1].height = required_row_height(
            heading_text, DASHBOARD_FULL_SPAN_PTS, 11,
        )

        pointer = (
            "This applies to every round-specific citation in this workbook, "
            f"including the '{Q25_BASIS_SHEET_NAME}' sheet."
        )
        ws.merge_cells("A2:F2")
        pointer_cell = ws.cell(row=2, column=1, value=pointer)
        pointer_cell.font = _font(color=xl_color("text_muted"), size=9)
        pointer_cell.alignment = _left()
        ws.row_dimensions[2].height = required_row_height(
            pointer, DASHBOARD_FULL_SPAN_PTS, 9,
        )

        for offset, para in enumerate(round_provenance_paragraphs()):
            row = 4 + offset
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value=para)
            cell.font = _font(size=9)
            cell.alignment = Alignment(horizontal="left", vertical="top",
                                       wrap_text=True)
            ws.row_dimensions[row].height = min(
                MAX_ROW_HEIGHT,
                required_row_height(para, DASHBOARD_FULL_SPAN_PTS, 9),
            )

    # ------------------------------------------------------------------
    # Generic DataFrame → worksheet writer
    # ------------------------------------------------------------------

    def _write_df_to_sheet(
        self,
        wb: "Workbook",
        sheet_name: str,
        df,
        title: str,
        currency_cols: list = None,
        pct_cols: list = None,
        number_cols: list = None,
        freeze_col: int = 2,
        max_rows: int = 200,
    ) -> None:
        """Write a DataFrame to a new worksheet with professional formatting."""
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

        if df is None or df.empty:
            ws["A1"] = "No data available."
            return

        display = df.head(max_rows)
        currency_cols = set(currency_cols or [])
        pct_cols = set(pct_cols or [])
        number_cols = set(number_cols or [])

        # FAIL LOUD ON A STALE CONFIG. Every one of these lists is a set of
        # column names typed at the call site, and five of the six sheets had
        # drifted: twenty names across Pipeline Detail, Distress Documentation,
        # Geographic Targeting, Impact Projections and Track Record referred to
        # columns that do not exist in the DataFrames they format. The columns
        # then fell through to the magnitude-based auto-detect below, which is
        # how Appendix C's share column ended up reading 0 0 0 0 0 0 1.
        #
        # A misconfigured column is silent by construction — the sheet still
        # renders, just wrong — so it has to be an exception rather than a
        # log line. The developer sees it on the first build; the CDE never
        # files it.
        named = currency_cols | pct_cols | number_cols
        missing = sorted(named - set(display.columns))
        if missing:
            raise ValueError(
                f"{sheet_name!r} formats column(s) that do not exist in its "
                f"table: {missing}. Available: {sorted(display.columns)}. "
                "A name that matches nothing formats nothing and falls through "
                "to magnitude-based auto-detection."
            )

        # Title row
        n_cols = len(display.columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = _font(bold=True, color="FFFFFF", size=12)
        title_cell.fill = _fill(xl_color("primary"))
        title_cell.alignment = _center()
        ws.row_dimensions[1].height = 24

        # Sub-header: generation info
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        sub_cell = ws.cell(row=2, column=1,
                           value=f"{self.application.cde.name}  |  {self.application.application_round}  "
                                 f"|  Generated {date.today().isoformat()}")
        sub_cell.font = _font(italic=True, color=xl_color("text_muted"), size=8)
        sub_cell.fill = _fill(xl_color("bg_light"))
        sub_cell.alignment = _center()
        ws.row_dimensions[2].height = 14

        # Header row (row 3)
        # WIDTHS BEFORE ROWS (1.3.0 B1, second site). Auto-sizing ran forty
        # lines BELOW the loop that set every row's height, so each height was
        # decided against a column width that did not exist yet — and the
        # heights were a hardcoded 16, uniformly, exactly as the dashboard's
        # metrics loop hardcoded 18. Same defect, six more sheets.
        #
        # Measured on the shipped sample: Investor Commitments!G6 holds the
        # 351-character "[CDE TO COMPLETE: This table is a blank form ...]"
        # instruction in a 40-wide column at height 16 — one line of nine, so
        # the half a CDE never saw included "Add one row per investor you can
        # name and defend to the CDFI Fund, and delete any row you cannot."
        # Distress Documentation!N8 clipped the ACS vintage off "CDFI Fund
        # NMTC Eligibility Table (2016-2020 ACS)".
        col_widths = {}
        for j, col_name in enumerate(display.columns, start=1):
            col_letter = get_column_letter(j)
            max_len = max(
                len(str(col_name)),
                max((len(str(v)) for v in display.iloc[:, j - 1] if v is not None), default=0),
            )
            width = min(max_len + 3, 40)
            ws.column_dimensions[col_letter].width = width
            col_widths[j] = width

        header_row = 3
        for j, col_name in enumerate(display.columns, start=1):
            cell = ws.cell(row=header_row, column=j, value=str(col_name))
            cell.font = _font(bold=True, color="FFFFFF", size=10)
            cell.fill = _fill(xl_color("primary"))
            cell.alignment = _center()
            cell.border = _thin_border()
        ws.row_dimensions[header_row].height = max(20.0, max(
            (required_row_height(str(c), span_points(col_widths[j]), 10)
             for j, c in enumerate(display.columns, start=1)),
            default=20.0,
        ))

        # Enable auto-filter on header row
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(n_cols)}{header_row + len(display)}"
        )

        # Freeze: row 3 header + first freeze_col columns
        ws.freeze_panes = ws.cell(row=header_row + 1, column=freeze_col + 1)

        # Data rows
        is_last_row = False
        for i, (_, row_data) in enumerate(display.iterrows()):
            row_num = header_row + 1 + i
            is_last_row = (i == len(display) - 1)
            bg = xl_color("row_total") if is_last_row else (
                xl_color("row_even") if i % 2 == 0 else xl_color("row_odd")
            )

            for j, (col_name, val) in enumerate(zip(display.columns, row_data), start=1):
                cell = ws.cell(row=row_num, column=j)
                cell.fill = _fill(bg)
                cell.border = _thin_border()
                cell.alignment = _left()

                if is_last_row:
                    cell.font = _font(bold=True, size=10)
                else:
                    cell.font = _font(size=10)

                # Format by column type
                if col_name in currency_cols and isinstance(val, (int, float)):
                    cell.value = val
                    cell.number_format = FMT_CURRENCY
                    cell.alignment = _right()
                elif col_name in pct_cols and isinstance(val, (int, float)):
                    cell.value = val
                    cell.number_format = FMT_PCT
                    cell.alignment = _right()
                # AN IDENTIFIER IS NOT A QUANTITY (1.2.1 B-3), AND HERE IT
                # BEATS BOTH THE number_cols CONFIG AND THE AUTO-DETECT.
                # "Award Year" is named in no list on the Track Record sheet,
                # so 2019 fell to the auto-detect below and took FMT_NUMBER —
                # which Excel displays as "2,019". That predates 1.2.1: the
                # workbook has rendered years with a thousands separator since
                # at least v1.2.0, and the three prose renderers joined it only
                # when 1.2.1 unified their formatting. Same rule, one place, so
                # a column added to number_cols by mistake still cannot put a
                # comma in a GEOID.
                elif is_identifier_column(col_name) and isinstance(val, (int, float)):
                    cell.value = val
                    cell.number_format = FMT_IDENTIFIER
                    cell.alignment = _left()
                elif col_name in number_cols and isinstance(val, (int, float)):
                    cell.value = val
                    cell.number_format = FMT_NUMBER
                    cell.alignment = _right()
                else:
                    # Auto-detect currency by value size for unlabelled columns
                    if isinstance(val, float) and abs(val) > 10000:
                        cell.value = val
                        cell.number_format = FMT_CURRENCY
                        cell.alignment = _right()
                    elif isinstance(val, (int, float)):
                        cell.value = val
                        cell.number_format = FMT_NUMBER
                        cell.alignment = _right()
                    else:
                        cell.value = str(val) if val is not None else ""

            # Derived from the widest thing this row actually has to display,
            # against the width its column actually got. 16 remains the floor,
            # so a row of short values looks exactly as it did.
            ws.row_dimensions[row_num].height = max(16.0, max(
                (required_row_height(ws.cell(row=row_num, column=j).value,
                                     span_points(col_widths[j]), 10)
                 for j in range(1, len(display.columns) + 1)),
                default=16.0,
            ))

        if len(df) > max_rows:
            note_row = header_row + max_rows + 2
            ws.cell(row=note_row, column=1,
                    value=f"Showing {max_rows} of {len(df)} rows. See full data in source system.")
            ws.cell(row=note_row, column=1).font = _font(italic=True, color=xl_color("text_muted"), size=9)

    # ------------------------------------------------------------------
    # Individual sheet builders
    # ------------------------------------------------------------------

    def _build_pipeline_sheet(self, wb: "Workbook") -> None:
        df = build_pipeline_table(self.application.pipeline, self.application.cde)
        self._write_df_to_sheet(
            wb, "Pipeline Detail", df,
            title="Appendix A: Pipeline Detail",
            currency_cols=["Total Project Cost ($)", "QEI Request ($)",
                           "Total QLICI ($)", "Leverage Loan ($)",
                           "Total NMTCs ($)", "Estimated Investor Equity ($)",
                           "CDE Fee ($)"],
            number_cols=["Jobs Created", "Jobs Retained",
                         "Affordable Units Built", "Square Feet"],
            freeze_col=2,
            max_rows=100,
        )

        # Conditional formatting: Distress Level column
        ws = wb["Pipeline Detail"]
        try:
            # Find "Distress Level" column index
            header_row_vals = [ws.cell(row=3, column=j).value for j in range(1, 50)]
            if "Distress Level" in header_row_vals:
                col_idx = header_row_vals.index("Distress Level") + 1
                col_letter = get_column_letter(col_idx)
                last_row = 3 + min(len(df), 100)
                # Deep Distress → green background
                green_fill = PatternFill(start_color=xl_color("green_light"),
                                        end_color=xl_color("green_light"), fill_type="solid")
                yellow_fill = PatternFill(start_color=xl_color("yellow_light"),
                                         end_color=xl_color("yellow_light"), fill_type="solid")
                red_fill = PatternFill(start_color=xl_color("red_light"),
                                       end_color=xl_color("red_light"), fill_type="solid")
                rng = f"{col_letter}4:{col_letter}{last_row}"
                ws.conditional_formatting.add(
                    rng, CellIsRule(operator="equal", formula=['"Deep Distress"'],
                                   fill=green_fill))
                ws.conditional_formatting.add(
                    rng, CellIsRule(operator="equal", formula=['"Severely Distressed"'],
                                   fill=yellow_fill))
                ws.conditional_formatting.add(
                    rng, CellIsRule(operator="equal", formula=['"Non-LIC (Ineligible)"'],
                                   fill=red_fill))
        except Exception:
            pass

    def _build_distress_sheet(self, wb: "Workbook") -> None:
        df = build_distress_table(self.application.pipeline)
        self._write_df_to_sheet(
            wb, "Distress Documentation", df,
            title="Appendix B: Distress Documentation",
            # No numeric columns. "Poverty Rate (%)" and "Unemployment Rate
            # (%)" both render the string "See ACS" — tables/distress_table
            # refuses to infer an ACS statistic from a distress label. The
            # pct_cols this used to carry named "Poverty Rate (Est.)" and
            # "Unemployment Rate (Est.)", which have never existed here.
            freeze_col=2,
        )

    def _build_geographic_sheet(self, wb: "Workbook") -> None:
        df = build_geographic_table(self.application.pipeline)
        self._write_df_to_sheet(
            wb, "Geographic Targeting", df,
            title="Appendix C: Geographic Targeting",
            currency_cols=["QEI ($)"],
            # The real pct_cols entry the 1.2.0 note asked for once this
            # config was repaired. The column is a float again.
            pct_cols=["QEI (% of Total)"],
            number_cols=["Project Count", "Deep/Severe Projects",
                         "Native Area Projects (CDE-declared)",
                         "HMR Projects", "OZ Projects"],
            freeze_col=1,
        )

        # Color scale on Total QEI column
        ws = wb["Geographic Targeting"]
        try:
            header_vals = [ws.cell(row=3, column=j).value for j in range(1, 30)]
            if "QEI ($)" in header_vals:
                col_idx = header_vals.index("QEI ($)") + 1
                col_letter = get_column_letter(col_idx)
                last_row = 3 + min(len(df), 200) if df is not None and not df.empty else 10
                ws.conditional_formatting.add(
                    f"{col_letter}4:{col_letter}{last_row}",
                    ColorScaleRule(
                        start_type="min", start_color="FFFFFF",
                        mid_type="percentile", mid_value=50, mid_color=xl_color("row_even"),
                        end_type="max", end_color=xl_color("secondary"),
                    )
                )
        except Exception:
            pass

    def _build_impact_sheet(self, wb: "Workbook") -> None:
        df = build_impact_table(self.application.pipeline)
        self._write_df_to_sheet(
            wb, "Impact Projections", df,
            title="Appendix D: Impact Projections",
            currency_cols=["QEI ($)", "Total Project Cost ($)",
                           "Cost per Job ($)", "QEI per Job ($)"],
            number_cols=["Jobs Created", "Jobs Retained", "Total Jobs",
                         "Affordable Units", "Commercial Sq Ft"],
            freeze_col=2,
            max_rows=100,
        )

    def _build_investor_sheet(self, wb: "Workbook") -> None:
        df = build_investor_table(self.application)
        self._write_df_to_sheet(
            wb, "Investor Commitments", df,
            title=INVESTOR_TABLE_TITLE,
            currency_cols=["Commitment Amount ($)", "NMTCs Allocated ($)"],
            freeze_col=1,
        )

    def _build_track_record_sheet(self, wb: "Workbook") -> None:
        df = build_track_record_table(self.application.cde)
        self._write_df_to_sheet(
            wb, "Track Record", df,
            title="Appendix E: CDE Track Record",
            currency_cols=["Allocation ($)"],
            freeze_col=1,
        )
